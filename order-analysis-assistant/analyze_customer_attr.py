#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
智能订单数据分析助手 - 客户属性维度分析

默认数据源：
- 区分客户自助下单_清洗版1.csv（必须含「客戶屬性」「创建时间」列）

功能：
1) 按客户属性分析各产品（小类）采购金额、箱数及占比
2) 按客户属性分析下单时段偏好（创建时间）

字段约定（兼容简繁体 / 常见别名）：
- 客户属性列：客户属性 / 客戶屬性  （维度筛选，必填）
- 金额列：订单总金额 / 訂單總金額（按订单分摊到行）
- 产品列：小类 / 小類
- 箱数列：销售数量
- 时间列：创建时间                 （时段分析必填）
"""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

COLUMN_ALIASES = {
    "客户属性": ["客户属性", "客戶屬性", "客戶性質", "客户性质"],
    "订单总金额": ["订单总金额", "訂單總金額"],
    "行金额": ["合计含税金额", "合計含稅金額", "金额", "金額"],
    "小类": ["小类", "小類", "产品", "產品", "商品小类"],
    "销售数量": ["销售数量", "銷售數量", "箱数", "箱數", "辅助数量"],
    "创建时间": ["创建时间", "創建時間", "创建時間", "下单时间", "下單時間"],
    "订单号": ["ERP订单号", "ERP訂單號", "订单号", "訂單號", "要货单号"],
}

ATTR_NORMALIZE = {
    "加工": "批發-加工",
    "飛機": "批發-飛機",
    "飞机": "批發-飛機",
    "烧腊": "燒臘",
    "肉台": "肉檯",
}

DEFAULT_CSV = str(
    Path(__file__).resolve().parent / "data" / "区分客户自助下单_清洗版1.csv"
)


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    return str(h).strip()


def resolve_columns(
    headers: Iterable[Any], require_time: bool = False
) -> Dict[str, str]:
    header_list = [_norm_header(h) for h in headers]
    mapping: Dict[str, str] = {}
    for canon, aliases in COLUMN_ALIASES.items():
        for h in header_list:
            if h in aliases:
                mapping[canon] = h
                break
    required = ["客户属性", "小类", "销售数量"]
    if require_time:
        required.append("创建时间")
    missing = [k for k in required if k not in mapping]
    if "订单总金额" not in mapping and "行金额" not in mapping:
        missing.append("订单总金额/行金额")
    if missing:
        raise ValueError(f"缺少必要列: {missing}; 实际表头={header_list}")
    return mapping


def normalize_attr_value(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if not s or s.upper() in {"#N/A", "N/A", "NONE", "NULL", "NAN"}:
        return ""
    return ATTR_NORMALIZE.get(s, s)


def normalize_source_rows(
    rows: List[Dict[str, Any]], col: Dict[str, str]
) -> List[Dict[str, Any]]:
    """清洗客戶屬性 / 创建时间，保证后续维度分析可用。"""
    attr_key = col["客户属性"]
    time_key = col.get("创建时间")
    cleaned: List[Dict[str, Any]] = []
    for r in rows:
        item = dict(r)
        item[attr_key] = normalize_attr_value(item.get(attr_key))
        if time_key:
            h = parse_hour(item.get(time_key))
            raw = item.get(time_key)
            if isinstance(raw, datetime):
                item[time_key] = raw.strftime("%Y-%m-%d %H:%M:%S")
            elif raw is None:
                item[time_key] = ""
            else:
                s = str(raw).strip()
                item[time_key] = s
            # 保留可解析小时的行标记，供校验
            item["__hour__"] = h
        cleaned.append(item)
    return cleaned


def list_customer_attrs(rows: List[Dict[str, Any]], col: Dict[str, str]) -> List[str]:
    cnt: Counter = Counter()
    key = col["客户属性"]
    for r in rows:
        v = normalize_attr_value(r.get(key))
        if v:
            cnt[v] += 1
    return [k for k, _ in cnt.most_common()]


def to_float(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return 0.0
        return float(v)
    s = str(v).strip().replace(",", "").replace("，", "")
    if not s or s.upper() in {"N/A", "#N/A", "NULL", "NONE"}:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def parse_hour(v: Any) -> Optional[int]:
    if v is None or v == "":
        return None
    if hasattr(v, "hour"):
        return int(v.hour)
    s = str(v).strip()
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y/%m/%d %H:%M",
        "%Y-%m-%dT%H:%M:%S",
        "%H:%M:%S",
        "%H:%M",
    ):
        try:
            return datetime.strptime(s, fmt).hour
        except ValueError:
            pass
    m = re.search(r"(\d{1,2}):\d{2}", s)
    if m:
        h = int(m.group(1))
        if 0 <= h <= 23:
            return h
    return None


def _order_key(row: Dict[str, Any], col: Dict[str, str], idx: int) -> str:
    if "订单号" in col:
        oid = row.get(col["订单号"])
        if oid not in (None, ""):
            return str(oid)
    return f"__row_{idx}"


def _line_amounts(rows: List[Dict[str, Any]], col: Dict[str, str]) -> List[float]:
    """优先使用可区分的行金额；否则按订单总金额依销售数量分摊。"""
    n = len(rows)
    amounts = [0.0] * n

    # 先看行金额是否在同订单内有差异
    if "行金额" in col:
        by_order: Dict[str, List[int]] = defaultdict(list)
        for i, r in enumerate(rows):
            by_order[_order_key(r, col, i)].append(i)
        use_line = False
        for idxs in by_order.values():
            vals = {round(to_float(rows[i].get(col["行金额"])), 4) for i in idxs}
            if len(idxs) > 1 and len(vals) > 1:
                use_line = True
                break
        if use_line:
            return [to_float(r.get(col["行金额"])) for r in rows]

    if "订单总金额" not in col:
        # 退化：直接用行金额
        if "行金额" in col:
            return [to_float(r.get(col["行金额"])) for r in rows]
        return amounts

    by_order = defaultdict(list)
    for i, r in enumerate(rows):
        by_order[_order_key(r, col, i)].append(i)

    for idxs in by_order.values():
        order_amt = 0.0
        for i in idxs:
            order_amt = max(order_amt, to_float(rows[i].get(col["订单总金额"])))
        qtys = [to_float(rows[i].get(col["销售数量"])) for i in idxs]
        qsum = sum(qtys)
        if qsum > 0:
            for i, q in zip(idxs, qtys):
                amounts[i] = order_amt * (q / qsum)
        else:
            share = order_amt / len(idxs)
            for i in idxs:
                amounts[i] = share
    return amounts


def filter_attr(rows: List[Dict[str, Any]], col: Dict[str, str], attr: str) -> List[Dict[str, Any]]:
    attr = normalize_attr_value(attr)
    out = []
    key = col["客户属性"]
    for r in rows:
        if normalize_attr_value(r.get(key)) == attr:
            out.append(r)
    return out


def analyze_sales_by_customer_attr(
    rows: List[Dict[str, Any]],
    headers: List[Any],
    customer_attr: str,
    top_n: int = 50,
) -> Dict[str, Any]:
    col = resolve_columns(headers, require_time=False)
    rows = normalize_source_rows(rows, col)
    scoped = filter_attr(rows, col, customer_attr)
    if not scoped:
        available = list_customer_attrs(rows, col)
        return {
            "analysis": "customer_attr_sales",
            "客户属性": customer_attr,
            "error": (
                f"未找到客户属性为「{customer_attr}」的订单数据。"
                f"当前「客戶屬性」可选值：{', '.join(available) if available else '无'}"
            ),
            "可用客户属性": available,
        }

    line_amts = _line_amounts(scoped, col)
    by_prod: Dict[str, List[float]] = defaultdict(lambda: [0.0, 0.0])
    for r, amt in zip(scoped, line_amts):
        prod = str(r.get(col["小类"]) or "未知").strip() or "未知"
        qty = to_float(r.get(col["销售数量"]))
        by_prod[prod][0] += amt
        by_prod[prod][1] += qty

    total_amt = sum(v[0] for v in by_prod.values())
    total_qty = sum(v[1] for v in by_prod.values())
    details = []
    for prod, (amt, qty) in by_prod.items():
        details.append(
            {
                "小类": prod,
                "金额": round(amt, 2),
                "箱数": round(qty, 2),
                "金额占比": round(amt / total_amt * 100, 2) if total_amt else 0.0,
                "箱数占比": round(qty / total_qty * 100, 2) if total_qty else 0.0,
            }
        )
    details.sort(key=lambda x: (-x["金额"], -x["箱数"], x["小类"]))
    if top_n and top_n > 0:
        details = details[:top_n]

    return {
        "analysis": "customer_attr_sales",
        "客户属性": normalize_attr_value(customer_attr),
        "维度字段": {"客户属性": col["客户属性"], "小类": col["小类"], "销售数量": col["销售数量"]},
        "订单行数": len(scoped),
        "产品数": len(by_prod),
        "总金额": round(total_amt, 2),
        "总箱数": round(total_qty, 2),
        "金额字段": col.get("订单总金额") or col.get("行金额"),
        "明细": details,
    }


def analyze_order_time_by_customer_attr(
    rows: List[Dict[str, Any]],
    headers: List[Any],
    customer_attr: str,
    window_hours: int = 2,
) -> Dict[str, Any]:
    col = resolve_columns(headers, require_time=True)
    rows = normalize_source_rows(rows, col)
    scoped = filter_attr(rows, col, customer_attr)
    if not scoped:
        available = list_customer_attrs(rows, col)
        return {
            "analysis": "customer_attr_order_time",
            "客户属性": customer_attr,
            "error": (
                f"未找到客户属性为「{customer_attr}」的订单数据。"
                f"当前「客戶屬性」可选值：{', '.join(available) if available else '无'}"
            ),
            "可用客户属性": available,
        }

    hour_cnt: Dict[int, int] = defaultdict(int)
    missing_time = 0
    for r in scoped:
        h = r.get("__hour__")
        if h is None:
            h = parse_hour(r.get(col["创建时间"]))
        if h is None:
            missing_time += 1
            continue
        hour_cnt[h] += 1
    total = sum(hour_cnt.values())
    if total == 0:
        return {
            "analysis": "customer_attr_order_time",
            "客户属性": customer_attr,
            "error": "「创建时间」列无法解析出有效小时，请检查清洗版 CSV 的创建时间格式",
            "创建时间缺失行数": missing_time,
        }

    window_hours = max(1, int(window_hours))
    best = (0, 0, window_hours)  # count, start, end_exclusive
    for start in range(0, 24 - window_hours + 1):
        c = sum(hour_cnt.get(h, 0) for h in range(start, start + window_hours))
        if c > best[0]:
            best = (c, start, start + window_hours)

    peak_h = max(hour_cnt, key=hour_cnt.get)
    hourly = []
    for h in range(24):
        c = hour_cnt.get(h, 0)
        if c:
            hourly.append(
                {
                    "小时": f"{h:02d}:00-{h+1:02d}:00",
                    "订单行数": c,
                    "占比": round(c / total * 100, 2),
                }
            )

    attr_norm = normalize_attr_value(customer_attr)
    return {
        "analysis": "customer_attr_order_time",
        "客户属性": attr_norm,
        "维度字段": {"客户属性": col["客户属性"], "创建时间": col["创建时间"]},
        "订单行数": total,
        "创建时间缺失行数": missing_time,
        "偏好时段": f"{best[1]:02d}:00-{best[2]:02d}:00",
        "偏好时段订单行数": best[0],
        "偏好时段占比": round(best[0] / total * 100, 2),
        "峰值小时": f"{peak_h:02d}:00-{peak_h+1:02d}:00",
        "峰值小时订单行数": hour_cnt[peak_h],
        "峰值小时占比": round(hour_cnt[peak_h] / total * 100, 2),
        "分时明细": hourly,
        "结论": (
            f"「{attr_norm}」更偏向于 {best[1]:02d}:00-{best[2]:02d}:00 下单，"
            f"该时段约占全部下单行的 {round(best[0] / total * 100, 2)}%；"
            f"峰值小时为 {peak_h:02d}:00-{peak_h+1:02d}:00。"
        ),
    }


def load_csv_rows(path: str) -> Tuple[List[str], List[Dict[str, Any]]]:
    with open(path, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            raise ValueError(f"CSV 无表头: {path}")
        headers = [_norm_header(h) for h in reader.fieldnames]
        rows: List[Dict[str, Any]] = []
        for raw in reader:
            row = {_norm_header(k): v for k, v in raw.items()}
            rows.append(row)
    return headers, rows


def load_excel_rows(path: str, sheet: Optional[str] = None) -> Tuple[List[str], List[Dict[str, Any]]]:
    try:
        import openpyxl
    except ImportError as e:
        raise ImportError("请先安装 openpyxl: pip install openpyxl") from e

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    headers: List[str] = []
    rows: List[Dict[str, Any]] = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [_norm_header(h) for h in row]
            continue
        rows.append(dict(zip(headers, row)))
    wb.close()
    return headers, rows


def load_order_rows(
    path: str, sheet: Optional[str] = None
) -> Tuple[List[str], List[Dict[str, Any]]]:
    """加载清洗版 CSV / Excel。优先识别 .csv。"""
    suffix = Path(path).suffix.lower()
    if suffix == ".csv":
        headers, rows = load_csv_rows(path)
    elif suffix in {".xlsx", ".xlsm", ".xls"}:
        headers, rows = load_excel_rows(path, sheet)
    else:
        # 尝试 CSV
        try:
            headers, rows = load_csv_rows(path)
        except Exception:
            headers, rows = load_excel_rows(path, sheet)

    # 强制校验关键维度列存在
    lower_map = {h: h for h in headers}
    has_attr = any(h in COLUMN_ALIASES["客户属性"] for h in headers)
    has_time = any(h in COLUMN_ALIASES["创建时间"] for h in headers)
    if not has_attr:
        raise ValueError(
            "数据源缺少「客戶屬性」列。请使用清洗版 CSV："
            "区分客户自助下单_清洗版1.csv"
        )
    if not has_time:
        raise ValueError(
            "数据源缺少「创建时间」列。请使用清洗版 CSV："
            "区分客户自助下单_清洗版1.csv"
        )
    _ = lower_map  # silence lint
    return headers, rows


def format_sales_report(result: Dict[str, Any]) -> str:
    if result.get("error"):
        return str(result["error"])
    lines = [
        f"## 客户属性销售分析：{result['客户属性']}",
        f"- 订单行数：{result['订单行数']}",
        f"- 产品数（小类）：{result['产品数']}",
        f"- 总金额：{result['总金额']:,.2f}",
        f"- 总箱数：{result['总箱数']:,.2f}",
        "",
        "| 小类 | 金额 | 金额占比 | 箱数 | 箱数占比 |",
        "|---|---:|---:|---:|---:|",
    ]
    for d in result["明细"]:
        lines.append(
            f"| {d['小类']} | {d['金额']:,.2f} | {d['金额占比']:.2f}% | {d['箱数']:,.2f} | {d['箱数占比']:.2f}% |"
        )
    return "\n".join(lines)


def format_time_report(result: Dict[str, Any]) -> str:
    if result.get("error"):
        return str(result["error"])
    lines = [
        f"## 客户属性下单时段分析：{result['客户属性']}",
        f"- 订单行数：{result['订单行数']}",
        f"- 偏好时段：{result['偏好时段']}（{result['偏好时段占比']:.2f}%）",
        f"- 峰值小时：{result['峰值小时']}（{result['峰值小时占比']:.2f}%）",
        f"- 结论：{result['结论']}",
        "",
        "| 时段 | 订单行数 | 占比 |",
        "|---|---:|---:|",
    ]
    for d in sorted(result["分时明细"], key=lambda x: -x["订单行数"]):
        lines.append(f"| {d['小时']} | {d['订单行数']} | {d['占比']:.2f}% |")
    return "\n".join(lines)


def main():
    parser = argparse.ArgumentParser(description="客户属性维度订单分析（清洗版 CSV）")
    parser.add_argument(
        "data",
        nargs="?",
        default=DEFAULT_CSV,
        help="订单数据路径（默认：data/区分客户自助下单_清洗版1.csv）",
    )
    parser.add_argument("--sheet", default=None, help="Excel 工作表名（CSV 可忽略）")
    parser.add_argument("--attr", default="", help="客户属性，如：凍肉店")
    parser.add_argument(
        "--mode",
        choices=["sales", "time", "both"],
        default="both",
        help="分析模式",
    )
    parser.add_argument("--window", type=int, default=2, help="偏好时段窗口小时数")
    parser.add_argument("--json", action="store_true", help="输出 JSON")
    parser.add_argument(
        "--list-attrs",
        action="store_true",
        help="仅列出数据源中的客戶屬性取值后退出",
    )
    args = parser.parse_args()

    headers, rows = load_order_rows(args.data, args.sheet)
    col = resolve_columns(headers, require_time=True)
    if args.list_attrs:
        print("\n".join(list_customer_attrs(rows, col)))
        return
    if not str(args.attr).strip():
        parser.error("请通过 --attr 指定客戶屬性，或使用 --list-attrs 查看可选值")

    outputs = []
    if args.mode in ("sales", "both"):
        outputs.append(analyze_sales_by_customer_attr(rows, headers, args.attr))
    if args.mode in ("time", "both"):
        outputs.append(
            analyze_order_time_by_customer_attr(rows, headers, args.attr, args.window)
        )

    if args.json:
        print(json.dumps(outputs if len(outputs) > 1 else outputs[0], ensure_ascii=False, indent=2))
        return

    for r in outputs:
        if r.get("analysis") == "customer_attr_sales":
            print(format_sales_report(r))
        else:
            print(format_time_report(r))
        print()


if __name__ == "__main__":
    main()
