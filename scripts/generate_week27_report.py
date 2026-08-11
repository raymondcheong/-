#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate 2026 Week 27 B-end mall operations analysis report."""

from __future__ import annotations

import base64
import io
import math
import re
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import font_manager
from matplotlib.colors import LinearSegmentedColormap, TwoSlopeNorm
from matplotlib.patches import Circle
from openpyxl import load_workbook
from scipy import stats
from scipy.stats import gaussian_kde
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler

ROOT = Path(__file__).resolve().parents[1]
REPORT_DIR = ROOT / "reports"
REPORT_DIR.mkdir(exist_ok=True)

for name in ["WenQuanYi Micro Hei", "Droid Sans Fallback"]:
    try:
        font_manager.findfont(name, fallback_to_default=False)
        plt.rcParams["font.sans-serif"] = [name]
        break
    except Exception:
        pass
plt.rcParams["axes.unicode_minus"] = False

BG = "#070b12"
CARD = "#0d1520"
CYAN = "#2ee6d6"
PINK = "#ff4f8b"
YELLOW = "#ffd166"
RED = "#ff5c5c"
TEXT = "#e8eef7"
MUTED = "#8b9bb4"


def fig_b64(fig, dpi: int = 150) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode()


def parse_fs(s) -> list[str]:
    return [a or b for a, b in re.findall(r"'([^']+)'|\"([^\"]+)\"", str(s or ""))]


def wow_html(cur: float, prev: float) -> str:
    if not prev:
        return ""
    r = (cur - prev) / prev
    arrow = "↑" if r >= 0 else "↓"
    cls = "up" if r >= 0 else "down"
    return f'<div class="wow {cls}">較上週 {arrow} {abs(r) * 100:.1f}%</div>'


def main() -> None:
    order_path = ROOT / "区分客户自助下单_20260521_006 (8).xlsx"
    multi_path = ROOT / "多元分析.xlsx"
    frozen_path = ROOT / "凍肉.xlsx"

    df = pd.read_excel(order_path, header=1)
    df = df.rename(
        columns={
            "客名": "customer",
            "销售数量": "qty_box",
            "合计含税金额": "amount",
            "ERP订单号": "erp",
            "ERP订单状态": "status",
            "创建时间": "created_at",
            "小類": "product",
            "客戶屬性": "attr",
        }
    )
    df["created_at"] = pd.to_datetime(df["created_at"])
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    df["qty_box"] = pd.to_numeric(df["qty_box"], errors="coerce")
    df = df[df["status"].astype(str).str.contains("登记|登記", na=False)].copy()
    df = df[df["erp"].notna() & (df["erp"].astype(str).str.strip() != "")].copy()

    week_start = pd.Timestamp("2026-06-29")
    week_end = pd.Timestamp("2026-07-05")
    as_of = pd.Timestamp("2026-07-06")

    week = df[(df["created_at"] >= week_start) & (df["created_at"] < week_end)].copy()
    hist = df[df["created_at"] < week_start].copy()
    all_hist = df.copy()
    week_orders = week.drop_duplicates("erp")

    n_orders = int(week_orders["erp"].nunique())
    total_amount = float(week_orders["amount"].sum())
    n_customers = int(week_orders["customer"].nunique())
    total_boxes = float(week["qty_box"].sum())

    push, clicks, devices = 8, 6, 270
    prev = dict(push=10, clicks=18, devices=262, orders=56, amount=140000, customers=18, boxes=311)
    prev_accum = 1_641_000
    target = 12_000_000
    accum = prev_accum + total_amount
    target_rate = accum / target

    attr_counts = week_orders["attr"].value_counts()
    prod = (
        week.groupby("product")
        .agg(boxes=("qty_box", "sum"), orders=("erp", "nunique"))
        .reset_index()
        .dropna(subset=["product"])
        .sort_values("boxes", ascending=False)
    )
    cust = (
        week.drop_duplicates("erp")
        .groupby("customer")
        .agg(orders=("erp", "nunique"), amount=("amount", "sum"))
        .reset_index()
        .sort_values(["orders", "amount"], ascending=[False, False])
    )
    new_customers = sorted(set(week_orders["customer"].dropna()) - set(hist["customer"].dropna()))

    # Association rules: yellow rows + same-category metrics
    wb = load_workbook(multi_path)
    ws = wb["product_level_rules"]
    same_df = pd.read_excel(multi_path, sheet_name="same_category_product_rules_wit")
    rules = []
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, 1)
        rgb = None
        try:
            rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
        except Exception:
            pass
        if not (rgb and "FFFF00" in str(rgb).upper()):
            continue
        ant, con = ws.cell(r, 1).value, ws.cell(r, 2).value
        ant_i, con_i = parse_fs(ant), parse_fs(con)
        extra: dict = {}
        for _, srow in same_df.iterrows():
            if set(parse_fs(srow["antecedents"])) == set(ant_i) and set(parse_fs(srow["consequents"])) == set(con_i):
                cis = srow["category_internal_support"]
                extra = {
                    "common_category": srow["common_category"],
                    "category_internal_support": None if pd.isna(cis) else float(cis),
                }
                break
        rules.append(
            {
                "antecedents": ant_i,
                "consequents": con_i,
                "support": float(ws.cell(r, 6).value),
                "confidence": float(ws.cell(r, 7).value),
                "lift": float(ws.cell(r, 8).value),
                **extra,
            }
        )

    # Residual: small category x attr from 凍肉.xlsx
    frozen = pd.read_excel(frozen_path, sheet_name="CRC_B15_OM_开单明细表（__060726凍肉").dropna(subset=["客戶屬性", "小类"])
    ct = pd.crosstab(frozen["客戶屬性"], frozen["小类"])
    _, _, _, expected = stats.chi2_contingency(ct)
    resid = (ct - expected) / np.sqrt(expected)
    top_cats = frozen["小类"].value_counts().head(18).index.tolist()
    attrs_order = [a for a in ["KA餐飲", "凍肉店", "加工", "小餐飲", "燒臘", "社會機構", "肉檯", "航空膳食", "酒店", "飛機"] if a in resid.index]
    resid_plot = resid.loc[attrs_order, [c for c in top_cats if c in resid.columns]].clip(lower=0)

    # Residual: hour x attr from self-order history
    ah_u = all_hist.dropna(subset=["attr", "created_at", "erp"]).drop_duplicates("erp").copy()
    ah_u["hour"] = ah_u["created_at"].dt.hour
    ct2 = pd.crosstab(ah_u["attr"], ah_u["hour"])
    _, _, _, exp2 = stats.chi2_contingency(ct2)
    resid2 = (ct2 - exp2) / np.sqrt(exp2)
    hours_show = [h for h in range(7, 23) if h in resid2.columns]
    attrs_t = [a for a in ["凍肉店", "加工", "燒臘", "肉檯", "飛機"] if a in resid2.index]
    rt = resid2.loc[attrs_t, hours_show]

    # K-means RFM
    cust_all = all_hist.drop_duplicates("erp")[["customer", "erp", "amount", "created_at"]]
    rfm = cust_all.groupby("customer").agg(
        last_date=("created_at", "max"),
        frequency=("erp", "nunique"),
        monetary=("amount", "sum"),
    ).reset_index()
    rfm["recency"] = (as_of - rfm["last_date"]).dt.days.clip(lower=0)
    rfm = rfm.dropna()
    x_log = np.column_stack(
        [rfm["recency"].values, np.log1p(rfm["frequency"].values), np.log1p(rfm["monetary"].values)]
    )
    xs = StandardScaler().fit_transform(x_log)
    rfm["cluster"] = KMeans(n_clusters=4, random_state=42, n_init=30).fit_predict(xs)
    label_names = ["超級VIP核心大客", "高價值活躍戶", "中等價值常規戶", "低價值沉睡戶"]
    raw_centers = []
    for i in range(4):
        sub = rfm[rfm["cluster"] == i]
        raw_centers.append((sub["monetary"].mean(), sub["frequency"].mean(), -sub["recency"].mean(), i))
    raw_centers.sort(reverse=True)
    order = [x[3] for x in raw_centers]
    cluster_map = {old: new for new, old in enumerate(order)}
    rfm["group"] = rfm["cluster"].map(cluster_map)
    groups = []
    for g in range(4):
        members = rfm[rfm["group"] == g].sort_values(["monetary", "frequency"], ascending=False)
        groups.append(
            {
                "name": label_names[g],
                "recency": float(members["recency"].mean()),
                "frequency": float(members["frequency"].mean()),
                "monetary": float(members["monetary"].mean()),
                "customers": members["customer"].tolist(),
            }
        )

    # Heatmap: all history, unique timestamps, Mon-Sat x 6-20
    hm = all_hist.dropna(subset=["created_at"]).drop_duplicates("created_at").copy()
    hm["dow"] = hm["created_at"].dt.dayofweek
    hm["hour"] = hm["created_at"].dt.hour
    hm = hm[(hm["dow"] <= 5) & (hm["hour"] >= 6) & (hm["hour"] <= 20)]
    heat = pd.crosstab(hm["dow"], hm["hour"]).reindex(index=range(6), columns=range(6, 21), fill_value=0)

    # Pareto
    cust_sorted = cust.sort_values("amount", ascending=False).reset_index(drop=True)
    n_top = max(1, int(math.ceil(len(cust_sorted) * 0.2)))
    top = cust_sorted.head(n_top)
    top_share = float(top["amount"].sum() / total_amount) if total_amount else 0.0
    pareto_labels = list(top["customer"]) + ["其他客戶"]
    pareto_vals = [float(x) for x in (top["amount"] / total_amount)] + [1 - top_share]

    # Amount distribution
    amt_hist = all_hist.drop_duplicates("erp")[["erp", "amount"]].dropna()
    bins = [0, 500, 1000, 2000, 5000, 10000, 20000, float("inf")]
    bin_labels = ["< 500", "500–999", "1k–1,999", "2k–4,999", "5k–9,999", "1萬–1.9萬", "≥ 2萬"]
    amt_hist["bucket"] = pd.cut(amt_hist["amount"], bins=bins, labels=bin_labels, right=False)
    hist_counts = amt_hist["bucket"].value_counts().reindex(bin_labels, fill_value=0)
    amounts = amt_hist["amount"].values.astype(float)
    amounts = amounts[np.isfinite(amounts)]
    kde = gaussian_kde(amounts)
    mids = np.array([250, 750, 1500, 3500, 7500, 15000, 30000], dtype=float)
    kde_raw = kde(mids)
    kde_scaled = kde_raw * (hist_counts.values.max() / max(kde_raw.max(), 1e-12))

    # ---- charts ----
    fig, ax = plt.subplots(figsize=(5.2, 4.2), subplot_kw=dict(polar=True), facecolor=BG)
    ax.set_facecolor(BG)
    cats = ["推送次數", "接收設備", "點擊次數", "下單次數", "下單箱數"]
    vals = np.array([push, devices, clicks, n_orders, total_boxes], dtype=float)
    refs = np.array([12, 300, 20, 70, 500], dtype=float)
    norm = np.clip(vals / refs, 0, 1)
    angles = np.linspace(0, 2 * np.pi, len(cats), endpoint=False).tolist() + [0]
    norm_c = norm.tolist() + [norm[0]]
    ax.plot(angles, norm_c, color=CYAN, linewidth=2)
    ax.fill(angles, norm_c, color=CYAN, alpha=0.25)
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(cats, color=TEXT, fontsize=9)
    ax.set_yticklabels([])
    ax.spines["polar"].set_color("#334")
    ax.grid(color="#334", alpha=0.6)
    for ang, v, n in zip(angles[:-1], vals, norm):
        ax.text(ang, min(n + 0.12, 1.05), f"{v:g}", color=CYAN, ha="center", fontsize=8)
    fig.suptitle("多維度相對表現", color=CYAN, fontsize=12, y=1.02)
    radar_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(10, 2.6), facecolor=BG)
    ax.set_facecolor(BG)
    ax.set_xlim(0, 10)
    ax.set_ylim(0, 3)
    ax.axis("off")
    nodes = [
        (1.2, "推送", push, CYAN),
        (3.6, "接收設備", devices, "#a78bfa"),
        (6.0, "點擊", clicks, PINK),
        (8.4, "下單", n_orders, "#4ade80"),
    ]
    for i, (x, lab, val, col) in enumerate(nodes):
        ax.add_patch(Circle((x, 1.5), 0.7, facecolor=col, alpha=0.9, edgecolor="white", linewidth=1.5))
        ax.text(x, 1.5, str(val), ha="center", va="center", fontsize=14, fontweight="bold", color="#041018")
        ax.text(x, 0.45, lab, ha="center", va="center", fontsize=11, color=TEXT)
        if i < len(nodes) - 1:
            ax.annotate(
                "",
                xy=(nodes[i + 1][0] - 0.75, 1.5),
                xytext=(x + 0.75, 1.5),
                arrowprops=dict(arrowstyle="->", color=MUTED, lw=2),
            )
    funnel_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(11, 4.8), facecolor=BG)
    ax.set_facecolor(BG)
    data = resid_plot.values.astype(float)
    cmap = LinearSegmentedColormap.from_list("or", ["#1a1208", "#f59e0b", "#ef4444", "#7f1d1d"])
    im = ax.imshow(data, aspect="auto", cmap=cmap)
    ax.set_xticks(range(len(resid_plot.columns)))
    ax.set_xticklabels(resid_plot.columns, rotation=45, ha="right", color=TEXT, fontsize=8)
    ax.set_yticks(range(len(resid_plot.index)))
    ax.set_yticklabels(resid_plot.index, color=TEXT, fontsize=9)
    for i in range(data.shape[0]):
        for j in range(data.shape[1]):
            if data[i, j] > 0.5:
                ax.text(j, i, f"{data[i, j]:.1f}", ha="center", va="center", color="white", fontsize=6)
    cb = fig.colorbar(im, ax=ax, fraction=0.025)
    cb.ax.yaxis.set_tick_params(color=TEXT)
    plt.setp(cb.ax.yaxis.get_ticklabels(), color=TEXT)
    ax.set_title("小類 × 客戶屬性 標準化殘差（正殘差＝偏好偏高）", color=CYAN, fontsize=12)
    resid_cat_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(11, 3.6), facecolor=BG)
    ax.set_facecolor(BG)
    vmin, vmax = float(rt.min().min()), float(rt.max().max())
    norm_rb = TwoSlopeNorm(vmin=vmin, vcenter=0, vmax=max(vmax, 0.1))
    cmap2 = LinearSegmentedColormap.from_list("rb", ["#1d4ed8", "#0b1220", "#ef4444"])
    im = ax.imshow(rt.values, aspect="auto", cmap=cmap2, norm=norm_rb)
    ax.set_xticks(range(len(hours_show)))
    ax.set_xticklabels(hours_show, color=TEXT)
    ax.set_yticks(range(len(attrs_t)))
    ax.set_yticklabels(attrs_t, color=TEXT)
    for i in range(rt.shape[0]):
        for j in range(rt.shape[1]):
            ax.text(j, i, f"{rt.values[i, j]:.1f}", ha="center", va="center", color="white", fontsize=7)
    cb = fig.colorbar(im, ax=ax, fraction=0.03)
    cb.ax.yaxis.set_tick_params(color=TEXT)
    plt.setp(cb.ax.yaxis.get_ticklabels(), color=TEXT)
    ax.set_title("下單時段 × 客戶屬性 標準化殘差（紅＝偏多／藍＝偏少）", color=CYAN, fontsize=12)
    resid_time_b64 = fig_b64(fig)

    labels_r, lifts = [], []
    for r in rules:
        labels_r.append("、".join(r["antecedents"]) + " → " + "、".join(r["consequents"]))
        lifts.append(r["lift"])
    order_idx = np.argsort(lifts)[::-1]
    labels_s = [labels_r[i] for i in order_idx]
    lifts_s = [lifts[i] for i in order_idx]
    fig, ax = plt.subplots(figsize=(10, 5.2), facecolor=BG)
    ax.set_facecolor(BG)
    y = np.arange(len(labels_s))
    ax.barh(y, lifts_s, color=CYAN, height=0.55)
    ax.set_yticks(y)
    ax.set_yticklabels(labels_s, color=TEXT, fontsize=9)
    ax.invert_yaxis()
    ax.tick_params(colors=TEXT)
    ax.set_xlabel("Lift（搭售提升倍數）", color=MUTED)
    for yi, v in zip(y, lifts_s):
        ax.text(v + 0.3, yi, f"{v:.2f} 倍", va="center", color=CYAN, fontsize=8)
    ax.set_title("搭售規則 · 售提升倍數對比", color=CYAN, fontsize=12)
    for spine in ax.spines.values():
        spine.set_color("#334")
    ax.grid(axis="x", color="#233", alpha=0.5)
    lift_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(11, 4.2), facecolor=BG)
    ax.set_facecolor(BG)
    cplot = cust.sort_values(["orders", "amount"], ascending=[False, False])
    x = np.arange(len(cplot))
    ax.bar(x, cplot["orders"], color=PINK, width=0.55, label="下單次數")
    ax2 = ax.twinx()
    ax2.plot(x, cplot["amount"] / 10000, color=YELLOW, marker="o", linewidth=2, label="銷售額(萬港元)")
    ax.set_xticks(x)
    ax.set_xticklabels(cplot["customer"], rotation=45, ha="right", color=TEXT, fontsize=8)
    ax.tick_params(colors=TEXT)
    ax2.tick_params(colors=TEXT)
    ax.set_ylabel("下單次數", color=PINK)
    ax2.set_ylabel("銷售額（萬港元）", color=YELLOW)
    ax.set_title("客戶下單次數與銷售額", color=CYAN, fontsize=12)
    for spine in list(ax.spines.values()) + list(ax2.spines.values()):
        spine.set_color("#334")
    cust_chart_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(11, 4.2), facecolor=BG)
    ax.set_facecolor(BG)
    pplot = prod.head(16)
    x = np.arange(len(pplot))
    ax.bar(x, pplot["boxes"], color=PINK, width=0.55)
    ax2 = ax.twinx()
    ax2.plot(x, pplot["orders"], color=YELLOW, marker="o", linewidth=2)
    ax.set_xticks(x)
    ax.set_xticklabels(pplot["product"], rotation=45, ha="right", color=TEXT, fontsize=8)
    ax.tick_params(colors=TEXT)
    ax2.tick_params(colors=TEXT)
    ax.set_ylabel("下單箱數", color=PINK)
    ax2.set_ylabel("品項下單次數", color=YELLOW)
    ax.set_title("產品下單統計", color=CYAN, fontsize=12)
    for spine in list(ax.spines.values()) + list(ax2.spines.values()):
        spine.set_color("#334")
    prod_chart_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(11, 3.8), facecolor=BG)
    ax.set_facecolor(BG)
    cmap_h = LinearSegmentedColormap.from_list("h", ["#fff7ed", "#fdba74", "#ea580c", "#7f1d1d"])
    im = ax.imshow(heat.values, aspect="auto", cmap=cmap_h)
    ax.set_xticks(range(15))
    ax.set_xticklabels(list(range(6, 21)), color=TEXT)
    ax.set_yticks(range(6))
    ax.set_yticklabels(["星期一", "星期二", "星期三", "星期四", "星期五", "星期六"], color=TEXT)
    for i in range(6):
        for j in range(15):
            v = int(heat.values[i, j])
            if v:
                ax.text(
                    j,
                    i,
                    str(v),
                    ha="center",
                    va="center",
                    color="#111" if v < heat.values.max() * 0.55 else "white",
                    fontsize=7,
                )
    cb = fig.colorbar(im, ax=ax, fraction=0.03)
    cb.set_label("訂單量", color=TEXT)
    cb.ax.yaxis.set_tick_params(color=TEXT)
    plt.setp(cb.ax.yaxis.get_ticklabels(), color=TEXT)
    ax.set_title("訂單量熱力分佈（週一至週六 × 06:00–20:00｜歷史全量｜同一時間計一次）", color=CYAN, fontsize=11)
    heat_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(6.2, 4.8), facecolor=BG)
    ax.set_facecolor(BG)
    colors = ["#22d3ee", "#a78bfa", "#f472b6", "#fbbf24", "#334155"]
    wedges, _, autotexts = ax.pie(
        pareto_vals,
        labels=None,
        colors=colors[: len(pareto_vals)],
        autopct=lambda p: f"{p:.1f}%" if p > 3 else "",
        pctdistance=0.75,
        wedgeprops=dict(width=0.42, edgecolor=BG, linewidth=2),
    )
    for t in autotexts:
        t.set_color("white")
        t.set_fontsize(9)
    ax.legend(
        wedges,
        [f"{l} {v * 100:.1f}%" for l, v in zip(pareto_labels, pareto_vals)],
        loc="center left",
        bbox_to_anchor=(1.0, 0.5),
        facecolor=CARD,
        edgecolor="#334",
        labelcolor=TEXT,
        fontsize=9,
    )
    ax.set_title("頭部客戶金額貢獻佔比", color=CYAN, fontsize=12)
    pareto_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(10, 4.2), facecolor=BG)
    ax.set_facecolor(BG)
    x = np.arange(len(bin_labels))
    ax.bar(x, hist_counts.values, color="#fb7185", width=0.65, label="訂單筆數")
    ax.plot(x, kde_scaled, color=RED, marker="o", linewidth=2, label="KDE")
    ax.set_xticks(x)
    ax.set_xticklabels(bin_labels, color=TEXT, fontsize=9)
    ax.tick_params(colors=TEXT)
    ax.set_ylabel("訂單筆數", color=TEXT)
    ax.set_title("客戶下單金額分布（直方圖 + 核密度估計）", color=CYAN, fontsize=12)
    ax.legend(facecolor=CARD, edgecolor="#334", labelcolor=TEXT)
    for spine in ax.spines.values():
        spine.set_color("#334")
    ax.grid(axis="y", color="#233", alpha=0.5)
    amt_b64 = fig_b64(fig)

    # insights
    flat = [(resid_plot.loc[a, c], a, c) for a in resid_plot.index for c in resid_plot.columns]
    flat.sort(reverse=True)
    insight_bits = [
        f"「{a}」對「{c}」偏好偏高（殘差 {v:.1f}）" for v, a, c in flat[:6] if v >= 2
    ]
    resid_insight = "；".join(insight_bits[:4]) if insight_bits else "本週殘差結構整體平穩"

    tflat = [(rt.loc[a, h], a, h) for a in rt.index for h in rt.columns]
    tflat.sort(reverse=True)
    time_parts = [f"「{a}」於 {h}:00 偏多（殘差 {v:.1f}）" for v, a, h in tflat[:5] if v >= 2]
    time_insight = "；".join(time_parts[:4]) if time_parts else "下單時段集中於上午"

    def rule_card(r: dict) -> str:
        ant = "、".join(r["antecedents"])
        con = "、".join(r["consequents"])
        extra = ""
        if r.get("common_category") is not None:
            extra += f'<div class="extra">common_category：{r["common_category"]}（同中類）</div>'
        if r.get("category_internal_support") is not None:
            extra += (
                f'<div class="extra">category_internal_support：'
                f'{r["category_internal_support"] * 100:.2f}%（同中類訂單內同時購買佔比）</div>'
            )
        return f"""<div class="rule-card card">
      <div class="rule-title">{ant} → {con}</div>
      <div class="metric">組合佔比（Support）：{r["support"] * 100:.2f}%（同單同時購買佔比）</div>
      <div class="metric">搭購轉化率（Confidence）：{r["confidence"] * 100:.1f}%（買了前者亦買後者）</div>
      <div class="metric">搭售提升倍數（Lift）：{r["lift"]:.2f} 倍（相對平常購買機率）</div>
      {extra}
    </div>"""

    attr_colors = {"飛機": "#38bdf8", "加工": "#a78bfa", "凍肉店": "#fb7185"}
    attr_cards = ""
    for a, c in attr_counts.items():
        share = c / n_orders
        attr_cards += f"""<div class="attr-card card"><div class="attr-name" style="color:{attr_colors.get(a, CYAN)}">{a}</div>
      <div class="attr-pct">{share * 100:.1f}%</div>
      <div class="attr-sub">{c}/{n_orders} 單</div></div>"""

    prod_detail = "、".join(
        [
            f"{r['product']}（{int(r['boxes']) if float(r['boxes']).is_integer() else r['boxes']} / {int(r['orders'])}）"
            for _, r in prod.iterrows()
        ]
    )

    group_html = ""
    for i, g in enumerate(groups, 1):
        names = "、".join(g["customers"]) if g["customers"] else "（無）"
        m = g["monetary"]
        m_txt = f"{m / 10000:.1f} 萬" if m >= 10000 else f"{m:,.2f}"
        group_html += f"""<div class="k-card card">
      <div class="k-title">第{i}組 · {g['name']}</div>
      <div class="k-metric">距今最近一次交易天數：<b>{g['recency']:.2f}</b></div>
      <div class="k-metric">交易頻次：<b>{g['frequency']:.2f}</b></div>
      <div class="k-metric">交易金額：<b>{m_txt}</b></div>
      <div class="k-list">客戶明細：{names}</div>
    </div>"""

    amt_table_rows = "".join(
        f"<tr><td>{lab}</td><td>{int(c)}</td><td>{k:.3f}</td></tr>"
        for lab, c, k in zip(bin_labels, hist_counts.values, kde_scaled)
    )
    new_cust_txt = "、".join(new_customers) if new_customers else "（無）"
    cust_delta = (
        "持平"
        if n_customers == prev["customers"]
        else f'變化 {(n_customers - prev["customers"]) / prev["customers"] * 100:+.1f}%'
    )

    html = f"""<!DOCTYPE html>
<html lang="zh-Hant">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>2026年第27周(6.29-7.4) B端商城運營分析報告</title>
<style>
:root {{
  --bg:#070b12; --card:#0d1520; --line:#1e2a3a; --cyan:#2ee6d6; --pink:#ff4f8b;
  --yellow:#ffd166; --text:#e8eef7; --muted:#8b9bb4;
}}
* {{ box-sizing:border-box; }}
body {{
  margin:0; font-family:"Microsoft JhengHei","PingFang TC","Noto Sans TC",sans-serif;
  background:radial-gradient(1200px 600px at 10% -10%, #122033 0%, var(--bg) 55%), var(--bg);
  color:var(--text); line-height:1.55;
}}
.wrap {{ max-width:1100px; margin:0 auto; padding:28px 20px 60px; }}
h1 {{ color:var(--cyan); font-size:28px; margin:0 0 8px; letter-spacing:.5px; }}
.sub {{ color:var(--muted); font-size:13px; margin-bottom:22px; }}
h2 {{ color:var(--cyan); font-size:20px; border-bottom:1px solid var(--line); padding-bottom:8px; margin:34px 0 14px; }}
.grid {{ display:grid; gap:12px; }}
.kpi-grid {{ grid-template-columns:repeat(4,1fr); }}
.card {{ background:linear-gradient(180deg, #101a27, #0b121c); border:1px solid var(--line); border-radius:14px; padding:14px 16px; }}
.kpi .label {{ color:var(--muted); font-size:12px; }}
.kpi .value {{ font-size:26px; font-weight:700; color:#fff; margin-top:4px; }}
.kpi .unit {{ font-size:12px; color:var(--muted); }}
.wow {{ font-size:12px; margin-top:6px; }}
.wow.up {{ color:#4ade80; }}
.wow.down {{ color:#fb7185; }}
.note {{ color:var(--muted); font-size:13px; margin:10px 0 0; }}
.editable {{ border:1px dashed #355; background:#0a1420; border-radius:10px; padding:10px 12px; min-height:42px; outline:none; }}
.editable:focus {{ border-color:var(--cyan); }}
.attr-grid {{ grid-template-columns:repeat(3,1fr); }}
.attr-card {{ text-align:center; padding:18px 10px; }}
.attr-name {{ font-size:18px; font-weight:700; }}
.attr-pct {{ font-size:32px; font-weight:800; margin:6px 0; }}
.attr-sub {{ color:var(--muted); font-size:13px; }}
.img {{ width:100%; border-radius:12px; border:1px solid var(--line); background:#05080e; }}
.rules {{ grid-template-columns:1fr 1fr; }}
.rule-title {{ font-weight:700; margin-bottom:8px; }}
.metric {{ color:var(--cyan); font-size:13px; margin:3px 0; }}
.extra {{ color:#86efac; font-size:12px; margin-top:4px; }}
.k-grid {{ grid-template-columns:repeat(4,1fr); }}
.k-card {{ min-height:220px; }}
.k-title {{ color:var(--cyan); font-weight:700; margin-bottom:8px; }}
.k-metric {{ font-size:13px; margin:4px 0; }}
.k-list {{ margin-top:8px; color:var(--muted); font-size:12px; word-break:break-all; }}
table {{ width:100%; border-collapse:collapse; font-size:13px; }}
th,td {{ border:1px solid var(--line); padding:8px 10px; text-align:left; }}
th {{ background:#121c2a; color:var(--cyan); }}
.summary-block {{ margin:10px 0 16px; }}
.summary-block h3 {{ margin:0 0 6px; color:#fff; font-size:15px; }}
.trad .value {{ font-size:18px; }}
@media (max-width:900px) {{
  .kpi-grid,.attr-grid,.rules,.k-grid {{ grid-template-columns:1fr 1fr; }}
}}
@media (max-width:600px) {{
  .kpi-grid,.attr-grid,.rules,.k-grid {{ grid-template-columns:1fr; }}
}}
</style>
</head>
<body>
<div class="wrap">
  <h1>2026年第27周(6.29-7.4) B端商城運營分析報告</h1>
  <div class="sub">推送效果 · 客戶轉化 · 產品下單 · 關聯規則 · 殘差與聚類 · 下週計劃　｜　v1 · week27_ops_report.html</div>

  <h2>一、核心指標總覽</h2>
  <div class="grid kpi-grid">
    <div class="card kpi"><div class="label">目標完成率</div><div class="value">{target_rate * 100:.2f}%</div><div class="unit">累計 {accum / 1e4:.1f} 萬 / 目標 {target / 1e4:.0f} 萬</div></div>
    <div class="card kpi"><div class="label">推送總次數</div><div class="value">{push}</div>{wow_html(push, prev["push"])}</div>
    <div class="card kpi"><div class="label">點擊總次數</div><div class="value">{clicks}</div>{wow_html(clicks, prev["clicks"])}</div>
    <div class="card kpi"><div class="label">成功接收設備數</div><div class="value">{devices}</div>{wow_html(devices, prev["devices"])}</div>
    <div class="card kpi"><div class="label">APP 下單次數</div><div class="value">{n_orders}</div>{wow_html(n_orders, prev["orders"])}<div class="unit">佔總下單次數比例（可編輯）</div><div class="editable" contenteditable="true">請手動填寫</div></div>
    <div class="card kpi"><div class="label">總金額（港元）</div><div class="value">{total_amount / 10000:.1f}萬</div>{wow_html(total_amount, prev["amount"])}<div class="unit">佔總下單金額（可編輯）</div><div class="editable" contenteditable="true">請手動填寫</div></div>
    <div class="card kpi trad"><div class="label">傳統渠道 APP 佔比（可編輯）</div><div class="value">金額佔比 / 次數佔比</div><div class="editable" contenteditable="true">金額：____%　｜　次數：____%</div></div>
    <div class="card kpi"><div class="label">下單箱數 / 下單客戶</div><div class="value">{int(total_boxes)} / {n_customers}</div>{wow_html(n_customers, prev["customers"])}</div>
  </div>
  <p class="note">本週篩選：R列時間 ∈ 2026-06-29～2026-07-04、P列＝已登記、O列非空；APP下單次數與總金額按 ERP 單號去重。產品名稱取 U列「小類」（T列為小類編碼），箱數取 I列。本週共 {n_orders} 單、{int(total_boxes)} 箱、{n_customers} 家客戶。</p>

  <h2>二、轉化流程網狀圖</h2>
  <img class="img" src="data:image/png;base64,{funnel_b64}" alt="轉化流程" />

  <h2>三、多維度雷達圖</h2>
  <img class="img" src="data:image/png;base64,{radar_b64}" alt="雷達圖" />

  <h2>四、客戶屬性下單佔比</h2>
  <div class="grid attr-grid">{attr_cards}</div>
  <p class="note">取數：V列客戶屬性；O列 ERP 單號去重後計次。</p>

  <h2>五、殘差分析</h2>
  <h3 style="color:#fff;font-size:15px;">小類 × 客戶屬性</h3>
  <img class="img" src="data:image/png;base64,{resid_cat_b64}" alt="小類殘差" />
  <p class="note">取數來源：凍肉.xlsx（客戶屬性 × 小类）。{resid_insight}。</p>
  <h3 style="color:#fff;font-size:15px;margin-top:18px;">下單時段 × 客戶屬性</h3>
  <img class="img" src="data:image/png;base64,{resid_time_b64}" alt="時段殘差" />
  <p class="note">取數來源：區分客戶自助下單表（歷史已登記訂單，O列去重）。{time_insight}。</p>

  <h2>六、K-means 客戶聚類分析</h2>
  <div class="grid k-grid">{group_html}</div>
  <p class="note">特徵：距今最近一次交易天數（R）、交易頻次（O唯一）、交易金額（L，同ERP不重複）；基於累計已登記訂單做 4 組聚類。</p>

  <h2>七、關聯規則 · 重點品項搭售組合</h2>
  <p class="note">本週依多元分析.xlsx 中 product_level_rules 標黃組合更新；同中類子表若有對應指標，則附加 common_category 與 category_internal_support。<br/>
  Support＝組合佔比；Confidence＝搭購轉化率；Lift＝相對平常購買機率的提升倍數（&gt;1 為正向關聯）。</p>
  <div class="grid rules">{"".join(rule_card(r) for r in rules)}</div>
  <img class="img" style="margin-top:12px" src="data:image/png;base64,{lift_b64}" alt="Lift對比" />
  <div class="card" style="margin-top:12px"><b>本週新增下單用戶（共 {len(new_customers)} 家）：</b>{new_cust_txt}</div>

  <h2>八、客戶下單次數與銷售額</h2>
  <p class="note">按客戶名稱彙總，共 {n_customers} 家；O列去重後 {n_orders} 單，總金額 {total_amount:,.0f} 港元（較上週客戶數 {prev["customers"]} 家 {cust_delta}）。</p>
  <img class="img" src="data:image/png;base64,{cust_chart_b64}" alt="客戶下單" />

  <h2>九、產品下單統計</h2>
  <p class="note">產品名稱取 U列「小類」，箱數加總 I列；本週合計 {int(total_boxes)} 箱。</p>
  <img class="img" src="data:image/png;base64,{prod_chart_b64}" alt="產品統計" />
  <p class="note">明細：{prod_detail}</p>

  <h2>十、圖一．訂單量熱力分佈（週一至週六 × 06:00–20:00）</h2>
  <p class="note">使用全部歷史已登記數據；R列時間，同一時間只計一次；細分至星期 × 小時。</p>
  <img class="img" src="data:image/png;base64,{heat_b64}" alt="熱力圖" />

  <h2>十一、圖二．客戶金額貢獻（頭部 20% 客戶／二八）</h2>
  <p class="note">本週按銷售額排序，頭部約 20% 客戶（{n_top} 家）貢獻總金額的 <b>{top_share * 100:.1f}%</b>；金額取 L列、客戶取 E列，同 ERP 不重複。</p>
  <img class="img" src="data:image/png;base64,{pareto_b64}" alt="二八" />
  <div class="card" style="margin-top:12px">
    <div style="color:var(--cyan);font-weight:700;margin-bottom:6px;">結論（可編輯）</div>
    <div class="editable" contenteditable="true">請在此書寫本週二八集中度結論……</div>
  </div>

  <h2>十二、圖五．客戶下單金額分布（直方圖＋核密度估計）</h2>
  <p class="note">歷史全量按 O列去重共 {len(amt_hist)} 筆訂單、{amt_hist["amount"].nunique()} 個不同金額（L列）；長條為頻次，紅線為 KDE。</p>
  <img class="img" src="data:image/png;base64,{amt_b64}" alt="金額分布" />
  <table style="margin-top:12px">
    <thead><tr><th>金額區間（港元）</th><th>訂單筆數</th><th>KDE</th></tr></thead>
    <tbody>{amt_table_rows}</tbody>
  </table>

  <h2>十三、本週整體情況分析</h2>
  <div class="summary-block card">
    <h3>運營情況</h3>
    <div class="editable" contenteditable="true">本週推送 {push} 次、成功接收設備 {devices} 個、點擊 {clicks} 次；APP 去重下單 {n_orders} 次。請補充運營側觀察與問題跟進……</div>
  </div>
  <div class="summary-block card">
    <h3>平台銷售額</h3>
    <div class="editable" contenteditable="true">本週 APP 下單金額 {total_amount:,.0f} 港元（約 {total_amount / 10000:.1f} 萬），較上週 {prev["amount"] / 10000:.0f} 萬 {(total_amount - prev["amount"]) / 10000:+.1f} 萬。頭部 20% 客戶貢獻 {top_share * 100:.1f}%。請補充銷售解讀……</div>
  </div>
  <div class="summary-block card">
    <h3>下週計劃</h3>
    <div class="editable" contenteditable="true">請填寫下週促銷節奏、重點客群與品項搭售實驗計劃……</div>
  </div>

</div>
</body>
</html>
"""

    out = REPORT_DIR / "2026_week27_ops_report.html"
    out.write_text(html, encoding="utf-8")
    print(f"Wrote {out} ({out.stat().st_size} bytes)")
    print(
        {
            "orders": n_orders,
            "amount": round(total_amount, 2),
            "customers": n_customers,
            "boxes": total_boxes,
            "new_customers": new_customers,
            "rules": len(rules),
            "groups": [(g["name"], len(g["customers"]), round(g["recency"], 2), round(g["frequency"], 2), round(g["monetary"], 2)) for g in groups],
            "pareto_share": round(top_share, 4),
        }
    )


if __name__ == "__main__":
    main()
