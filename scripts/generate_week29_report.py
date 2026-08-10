#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Generate 2026 Week 29 B-end mall operations analysis report.

Template / style / analysis framework mirror week 27/28 enhanced HTML report.
Week window: 2026-07-13 ~ 2026-07-18 (R column).
"""

from __future__ import annotations

import base64
import io
import json
import math
import re
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from matplotlib import font_manager
from matplotlib.colors import LinearSegmentedColormap, TwoSlopeNorm
from matplotlib.patches import Circle
from openpyxl import load_workbook
from scipy import stats
from scipy.stats import gaussian_kde
from sklearn.cluster import KMeans
from sklearn.preprocessing import StandardScaler

ROOT = Path(__file__).resolve().parents[1]
REPORT_DIR = ROOT / "reports"
REPORT_DIR.mkdir(exist_ok=True)

for name in ["WenQuanYi Micro Hei", "Droid Sans Fallback"]:
    try:
        font_manager.findfont(name, fallback_to_default=False)
        plt.rcParams["font.sans-serif"] = [name]
        break
    except Exception:
        pass
plt.rcParams["axes.unicode_minus"] = False

BG = "#070b12"
CARD = "#0d1520"
CYAN = "#2ee6d6"
PINK = "#ff4f8b"
YELLOW = "#ffd166"
RED = "#ff5c5c"
TEXT = "#e8eef7"
MUTED = "#8b9bb4"


def fig_b64(fig, dpi: int = 150) -> str:
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=dpi, bbox_inches="tight", facecolor=fig.get_facecolor())
    plt.close(fig)
    return base64.b64encode(buf.getvalue()).decode()


def parse_fs(s) -> list[str]:
    return [a or b for a, b in re.findall(r"'([^']+)'|\"([^\"]+)\"", str(s or ""))]


def wow_html(cur: float, prev: float) -> str:
    if not prev:
        return ""
    r = (cur - prev) / prev
    arrow = "↑" if r >= 0 else "↓"
    cls = "up" if r >= 0 else "down"
    return f'<div class="wow {cls}">較上週 {arrow} {abs(r) * 100:.1f}%</div>'


def resolve_order_file() -> Path:
    candidates = [
        ROOT / "区分客户自助下单_20260521_006_第十九階段.xlsx",
        ROOT / "区分客户自助下单_20260521_006 (9).xlsx",
        ROOT / "区分客户自助下单_20260521_006 (8).xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError("找不到自助下單 Excel（第十九階段 / (9) / (8)）")


def resolve_rules_file() -> tuple[Path, str]:
    """Return (path, mode) identifying which rules workbook was selected."""
    candidates = [
        (ROOT / "product_level_rules .xlsx", "product_level_rules"),
        (ROOT / "product_level_rules.xlsx", "product_level_rules"),
        (ROOT / "product_level_rules(18).xlsx", "product_level_rules(18)"),
        (ROOT / "多元分析.xlsx", "多元分析"),
    ]
    for path, mode in candidates:
        if path.exists():
            return path, mode
    raise FileNotFoundError(
        "找不到 product_level_rules .xlsx / product_level_rules.xlsx / 多元分析.xlsx"
    )


def resolve_frozen_file() -> Path:
    candidates = [
        ROOT / "凍肉_第十九階段.xlsx",
        ROOT / "凍肉(18).xlsx",
        ROOT / "凍肉.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError("找不到 凍肉_第十九階段.xlsx / 凍肉(18).xlsx / 凍肉.xlsx")


def normalize_attr(val) -> str | None:
    """Normalize customer attribute labels to report vocabulary."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return None
    if val is False:
        return None
    s = str(val).strip()
    if not s or s.lower() in {"nan", "none", "false"}:
        return None
    mapping = {
        "批發-加工": "加工",
        "批發-飛機": "飛機",
        "小餐": "小餐飲",
        "加工": "加工",
        "飛機": "飛機",
        "凍肉店": "凍肉店",
        "燒臘": "燒臘",
        "肉檯": "肉檯",
        "小餐飲": "小餐飲",
        "KA餐飲": "KA餐飲",
        "酒店": "酒店",
        "航空膳食": "航空膳食",
        "社會機構": "社會機構",
    }
    return mapping.get(s, s)


def load_attr_lookup(order_path: Path) -> tuple[dict, dict]:
    """Load 客戶屬性1 sheet: code→attr and 客名→attr.

    Stage-19 order exports store broken VLOOKUP results in the 客戶屬性 column
    (company names). Prefer the workbook's 客戶屬性1 mapping sheet.
    """
    code_map: dict = {}
    name_map: dict = {}
    try:
        wb = load_workbook(order_path, data_only=False, read_only=True)
    except Exception:
        return code_map, name_map
    if "客戶屬性1" not in wb.sheetnames:
        wb.close()
        return code_map, name_map
    ws = wb["客戶屬性1"]
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if not header:
        wb.close()
        return code_map, name_map
    # Expected: 客户账号, 客户名称, 客戶性質, ..., 客戶編碼, 客名
    h = [str(x).strip() if x is not None else "" for x in header]
    def idx(*names):
        for n in names:
            if n in h:
                return h.index(n)
        return None

    i_attr = idx("客戶性質", "客戶屬性")
    i_code = idx("客戶編碼", "客户编码")
    i_short = idx("客名")
    i_full = idx("客户名称", "客戶名稱")
    for row in rows:
        if not row:
            continue
        attr = normalize_attr(row[i_attr]) if i_attr is not None and i_attr < len(row) else None
        if not attr:
            continue
        if i_code is not None and i_code < len(row) and row[i_code]:
            code_map[str(row[i_code]).strip()] = attr
        if i_short is not None and i_short < len(row) and row[i_short]:
            name_map[str(row[i_short]).strip()] = attr
        if i_full is not None and i_full < len(row) and row[i_full]:
            name_map[str(row[i_full]).strip()] = attr
    wb.close()
    return code_map, name_map


def main() -> None:
    order_path = resolve_order_file()
    rules_path, rules_mode = resolve_rules_file()
    frozen_path = resolve_frozen_file()

    df = pd.read_excel(order_path, header=1)
    # Keep original 客戶編碼 / 客户 for attribute lookup before rename
    if "客戶編碼" not in df.columns and "客户编码" in df.columns:
        df = df.rename(columns={"客户编码": "客戶編碼"})
    code_map, name_map = load_attr_lookup(order_path)

    df = df.rename(
        columns={
            "客名": "customer",
            "销售数量": "qty_box",
            "合计含税金额": "amount",
            "ERP订单号": "erp",
            "ERP订单状态": "status",
            "创建时间": "created_at",
            "小類": "product",
            "客戶屬性": "attr_raw",
        }
    )
    # Resolve attr: workbook 客戶屬性1 (by code / 客名) → normalized raw column
    def resolve_row_attr(row) -> str | None:
        code = str(row.get("客戶編碼", "") or "").strip()
        cust = str(row.get("customer", "") or "").strip()
        full = str(row.get("客户", "") or "").strip()
        for key in (code, cust, full):
            if key and key in code_map:
                return code_map[key]
            if key and key in name_map:
                return name_map[key]
        return normalize_attr(row.get("attr_raw"))

    df["created_at"] = pd.to_datetime(df["created_at"])
    df["amount"] = pd.to_numeric(df["amount"], errors="coerce")
    df["qty_box"] = pd.to_numeric(df["qty_box"], errors="coerce")
    df = df[df["status"].astype(str).str.contains("登记|登記", na=False)].copy()
    df = df[df["erp"].notna() & (df["erp"].astype(str).str.strip() != "")].copy()
    df["attr"] = df.apply(resolve_row_attr, axis=1)

    # Week 29: 13/7/2026 ~ 18/7/2026 (D/M/Y)
    week_start = pd.Timestamp("2026-07-13")
    week_end = pd.Timestamp("2026-07-19")  # exclusive
    as_of = pd.Timestamp("2026-07-20")

    week = df[(df["created_at"] >= week_start) & (df["created_at"] < week_end)].copy()
    hist = df[df["created_at"] < week_start].copy()
    all_hist = df.copy()
    week_orders = week.drop_duplicates("erp")

    data_note = ""
    if "第十九階段" not in order_path.name:
        data_note = (
            f"注意：工作區未找到「区分客户自助下单_20260521_006_第十九階段.xlsx」，"
            f"已改用「{order_path.name}」。該檔最新時間為 {df['created_at'].max()}。"
        )

    n_orders = int(week_orders["erp"].nunique())
    total_amount = float(week_orders["amount"].sum()) if n_orders else 0.0
    n_customers = int(week_orders["customer"].nunique()) if n_orders else 0
    total_boxes = float(week["qty_box"].sum()) if len(week) else 0.0

    # Push metrics provided by user for week 29
    push, clicks, devices = 10, 10, 276
    # Week 28 baselines
    prev = dict(
        push=8,
        clicks=6,
        devices=270,
        orders=49,
        amount=131397.43,
        customers=19,
        boxes=259,
    )
    prev_accum = 1_958_865.99  # 1,827,468.56 + 131,397.43
    target = 12_000_000
    accum = prev_accum + total_amount
    target_rate = accum / target

    attr_counts = week_orders["attr"].value_counts() if n_orders else pd.Series(dtype=int)
    prod = (
        week.groupby("product")
        .agg(boxes=("qty_box", "sum"), orders=("erp", "nunique"))
        .reset_index()
        .dropna(subset=["product"])
        .sort_values("boxes", ascending=False)
    )
    cust = (
        week.drop_duplicates("erp")
        .groupby("customer")
        .agg(orders=("erp", "nunique"), amount=("amount", "sum"))
        .reset_index()
        .sort_values(["orders", "amount"], ascending=[False, False])
    )
    new_customers = sorted(set(week_orders["customer"].dropna()) - set(hist["customer"].dropna()))

    # Association rules: yellow rows + same-category metrics
    wb = load_workbook(rules_path)
    sheet_names = wb.sheetnames
    rules_sheet = sheet_names[0]
    for cand in sheet_names:
        name = str(cand).strip()
        if "product_level_rules" in name and "同中類" not in name and "same_category" not in name:
            rules_sheet = cand
            break
    ws = wb[rules_sheet]
    same_name = None
    for cand in ["同中類", "same_category_product_rules_wit", "same_category"]:
        # exact or stripped match
        if cand in sheet_names:
            same_name = cand
            break
        for sn in sheet_names:
            if str(sn).strip() == cand or cand in str(sn):
                same_name = sn
                break
        if same_name:
            break
    same_df = pd.read_excel(rules_path, sheet_name=same_name) if same_name else pd.DataFrame()

    # Detect columns on main rules sheet; still join 同中類 for common_category when present
    header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col_map = {str(h).strip(): i + 1 for i, h in enumerate(header_row) if h}

    rules = []
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(r, 1)
        rgb = None
        try:
            rgb = cell.fill.fgColor.rgb if cell.fill.fgColor else None
        except Exception:
            pass
        # Also accept theme/indexed yellow approximations
        is_yellow = bool(rgb and "FFFF00" in str(rgb).upper())
        if not is_yellow:
            # fallback: check patternType fill start color
            try:
                fc = cell.fill.start_color.rgb if cell.fill.start_color else None
                is_yellow = bool(fc and "FFFF00" in str(fc).upper())
            except Exception:
                pass
        if not is_yellow:
            continue

        ant_col = col_map.get("antecedents", 1)
        con_col = col_map.get("consequents", 2)
        ant, con = ws.cell(r, ant_col).value, ws.cell(r, con_col).value
        ant_i, con_i = parse_fs(ant), parse_fs(con)
        support = float(ws.cell(r, col_map.get("support", 5)).value)
        confidence = float(ws.cell(r, col_map.get("confidence", 6)).value)
        lift = float(ws.cell(r, col_map.get("lift", 7)).value)
        extra: dict = {}

        # Prefer metrics already on the main sheet when present
        if "common_category" in col_map:
            cc = ws.cell(r, col_map["common_category"]).value
            if cc is not None and str(cc).strip() != "":
                extra["common_category"] = cc
        if "category_internal_support" in col_map:
            cis = ws.cell(r, col_map["category_internal_support"]).value
            if cis is not None and str(cis).strip() != "" and not (isinstance(cis, float) and np.isnan(cis)):
                extra["category_internal_support"] = float(cis)

        # Always try 同中類 / same_category sheet to fill common_category
        # (and category_internal_support if missing on the main sheet)
        if not same_df.empty and (
            "common_category" not in extra or "category_internal_support" not in extra
        ):
            for _, srow in same_df.iterrows():
                if set(parse_fs(srow.get("antecedents"))) == set(ant_i) and set(
                    parse_fs(srow.get("consequents"))
                ) == set(con_i):
                    if "common_category" not in extra and "common_category" in srow and pd.notna(
                        srow["common_category"]
                    ):
                        extra["common_category"] = srow["common_category"]
                    if (
                        "category_internal_support" not in extra
                        and "category_internal_support" in srow
                        and pd.notna(srow["category_internal_support"])
                    ):
                        extra["category_internal_support"] = float(srow["category_internal_support"])
                    break

        rules.append(
            {
                "antecedents": ant_i,
                "consequents": con_i,
                "support": support,
                "confidence": confidence,
                "lift": lift,
                **extra,
            }
        )
    wb.close()

    # Residual: small category x attr from 凍肉.xlsx
    frozen_sheets = pd.ExcelFile(frozen_path).sheet_names
    frozen_sheet = [s for s in frozen_sheets if "开单" in s or "開單" in s or "凍肉" in s]
    frozen_sheet = frozen_sheet[0] if frozen_sheet else frozen_sheets[-1]
    frozen = pd.read_excel(frozen_path, sheet_name=frozen_sheet)
    # tolerate column name variants
    attr_col = "客戶屬性" if "客戶屬性" in frozen.columns else [c for c in frozen.columns if "屬性" in str(c) or "性質" in str(c)][0]
    cat_col = "小类" if "小类" in frozen.columns else ("小類" if "小類" in frozen.columns else None)
    if cat_col is None:
        raise KeyError(f"凍肉.xlsx 找不到小类列，現有：{list(frozen.columns)}")
    frozen = frozen.dropna(subset=[attr_col, cat_col]).copy()
    frozen[attr_col] = frozen[attr_col].map(normalize_attr)
    frozen = frozen.dropna(subset=[attr_col, cat_col])
    ct = pd.crosstab(frozen[attr_col], frozen[cat_col])
    _, _, _, expected = stats.chi2_contingency(ct)
    resid = (ct - expected) / np.sqrt(expected)
    attrs_order = [
        a
        for a in ["KA餐飲", "凍肉店", "加工", "小餐飲", "燒臘", "社會機構", "肉檯", "航空膳食", "酒店", "飛機"]
        if a in resid.index
    ]
    # full residual matrix for Plotly (all categories), signed
    resid_full = resid.loc[attrs_order]
    # static PNG: top categories, positive only (legacy look)
    top_cats = frozen[cat_col].value_counts().head(18).index.tolist()
    resid_plot = resid.loc[attrs_order, [c for c in top_cats if c in resid.columns]].clip(lower=0)

    # Residual: hour x attr from self-order history
    ah_u = all_hist.dropna(subset=["attr", "created_at", "erp"]).drop_duplicates("erp").copy()
    ah_u["hour"] = ah_u["created_at"].dt.hour
    ct2 = pd.crosstab(ah_u["attr"], ah_u["hour"])
    _, _, _, exp2 = stats.chi2_contingency(ct2)
    resid2 = (ct2 - exp2) / np.sqrt(exp2)
    hours_show = [h for h in range(7, 23) if h in resid2.columns]
    attrs_t = [a for a in ["凍肉店", "加工", "燒臘", "肉檯", "飛機", "小餐飲", "酒店"] if a in resid2.index]
    if not attrs_t:
        attrs_t = list(resid2.index)[:5]
    rt = resid2.loc[attrs_t, hours_show] if attrs_t and hours_show else resid2.iloc[:1, :1]

    # K-means RFM on all historical registered orders
    cust_all = all_hist.drop_duplicates("erp")[["customer", "erp", "amount", "created_at"]]
    rfm = (
        cust_all.groupby("customer")
        .agg(
            last_date=("created_at", "max"),
            frequency=("erp", "nunique"),
            monetary=("amount", "sum"),
        )
        .reset_index()
    )
    rfm["recency"] = (as_of - rfm["last_date"]).dt.days.clip(lower=0)
    rfm = rfm.dropna()
    x_log = np.column_stack(
        [rfm["recency"].values, np.log1p(rfm["frequency"].values), np.log1p(rfm["monetary"].values)]
    )
    xs = StandardScaler().fit_transform(x_log)
    rfm["cluster"] = KMeans(n_clusters=4, random_state=42, n_init=30).fit_predict(xs)
    label_names = ["超級VIP核心大客", "高價值活躍戶", "中等價值常規戶", "低價值沉睡戶"]
    raw_centers = []
    for i in range(4):
        sub = rfm[rfm["cluster"] == i]
        raw_centers.append((sub["monetary"].mean(), sub["frequency"].mean(), -sub["recency"].mean(), i))
    raw_centers.sort(reverse=True)
    order = [x[3] for x in raw_centers]
    cluster_map = {old: new for new, old in enumerate(order)}
    rfm["group"] = rfm["cluster"].map(cluster_map)
    groups = []
    for g in range(4):
        members = rfm[rfm["group"] == g].sort_values(["monetary", "frequency"], ascending=False)
        groups.append(
            {
                "name": label_names[g],
                "recency": float(members["recency"].mean()) if len(members) else 0.0,
                "frequency": float(members["frequency"].mean()) if len(members) else 0.0,
                "monetary": float(members["monetary"].mean()) if len(members) else 0.0,
                "customers": members["customer"].tolist(),
            }
        )

    # Heatmap: all history, unique timestamps, Mon-Sat x 6-20
    hm = all_hist.dropna(subset=["created_at"]).drop_duplicates("created_at").copy()
    hm["dow"] = hm["created_at"].dt.dayofweek
    hm["hour"] = hm["created_at"].dt.hour
    hm = hm[(hm["dow"] <= 5) & (hm["hour"] >= 6) & (hm["hour"] <= 20)]
    heat = pd.crosstab(hm["dow"], hm["hour"]).reindex(index=range(6), columns=range(6, 21), fill_value=0)

    # Pareto (week)
    if n_orders and total_amount > 0:
        cust_sorted = cust.sort_values("amount", ascending=False).reset_index(drop=True)
        n_top = max(1, int(math.ceil(len(cust_sorted) * 0.2)))
        top = cust_sorted.head(n_top)
        top_share = float(top["amount"].sum() / total_amount)
        pareto_labels = list(top["customer"]) + ["其他客戶"]
        pareto_vals = [float(x) for x in (top["amount"] / total_amount)] + [max(0.0, 1 - top_share)]
    else:
        n_top = 0
        top_share = 0.0
        pareto_labels = ["無本週訂單"]
        pareto_vals = [1.0]

    # Amount distribution (all history, unique ERP)
    amt_hist = all_hist.drop_duplicates("erp")[["erp", "amount"]].dropna()
    bins = [0, 500, 1000, 2000, 5000, 10000, 20000, float("inf")]
    bin_labels = ["< 500", "500–999", "1k–1,999", "2k–4,999", "5k–9,999", "1萬–1.9萬", "≥ 2萬"]
    amt_hist["bucket"] = pd.cut(amt_hist["amount"], bins=bins, labels=bin_labels, right=False)
    hist_counts = amt_hist["bucket"].value_counts().reindex(bin_labels, fill_value=0)
    amounts = amt_hist["amount"].values.astype(float)
    amounts = amounts[np.isfinite(amounts)]
    kde = gaussian_kde(amounts) if len(amounts) >= 2 else None
    mids = np.array([250, 750, 1500, 3500, 7500, 15000, 30000], dtype=float)
    if kde is not None:
        kde_raw = kde(mids)
        kde_scaled = kde_raw * (hist_counts.values.max() / max(kde_raw.max(), 1e-12))
    else:
        kde_scaled = np.zeros(len(bin_labels))

    # ---- charts ----
    fig, ax = plt.subplots(figsize=(5.2, 4.2), subplot_kw=dict(polar=True), facecolor=BG)
    ax.set_facecolor(BG)
    cats = ["推送次數", "接收設備", "點擊次數", "下單次數", "下單箱數"]
    vals = np.array([push, devices, clicks, n_orders, total_boxes], dtype=float)
    refs = np.array([12, 300, 20, 70, 500], dtype=float)
    norm = np.clip(vals / refs, 0, 1)
    angles = np.linspace(0, 2 * np.pi, len(cats), endpoint=False).tolist() + [0]
    norm_c = norm.tolist() + [norm[0]]
    ax.plot(angles, norm_c, color=CYAN, linewidth=2)
    ax.fill(angles, norm_c, color=CYAN, alpha=0.25)
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels(cats, color=TEXT, fontsize=9)
    ax.set_yticklabels([])
    ax.spines["polar"].set_color("#334")
    ax.grid(color="#334", alpha=0.6)
    for ang, v, n in zip(angles[:-1], vals, norm):
        ax.text(ang, min(n + 0.12, 1.05), f"{v:g}", color=CYAN, ha="center", fontsize=8)
    fig.suptitle("多維度相對表現", color=CYAN, fontsize=12, y=1.02)
    radar_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(10, 2.6), facecolor=BG)
    ax.set_facecolor(BG)
    ax.set_xlim(0, 10)
    ax.set_ylim(0, 3)
    ax.axis("off")
    nodes = [
        (1.2, "推送", push, CYAN),
        (3.6, "接收設備", devices, "#a78bfa"),
        (6.0, "點擊", clicks, PINK),
        (8.4, "下單", n_orders, "#4ade80"),
    ]
    for i, (x, lab, val, col) in enumerate(nodes):
        ax.add_patch(Circle((x, 1.5), 0.7, facecolor=col, alpha=0.9, edgecolor="white", linewidth=1.5))
        ax.text(x, 1.5, str(val), ha="center", va="center", fontsize=14, fontweight="bold", color="#041018")
        ax.text(x, 0.45, lab, ha="center", va="center", fontsize=11, color=TEXT)
        if i < len(nodes) - 1:
            ax.annotate(
                "",
                xy=(nodes[i + 1][0] - 0.75, 1.5),
                xytext=(x + 0.75, 1.5),
                arrowprops=dict(arrowstyle="->", color=MUTED, lw=2),
            )
    funnel_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(11, 3.6), facecolor=BG)
    ax.set_facecolor(BG)
    vmin, vmax = float(rt.min().min()), float(rt.max().max())
    norm_rb = TwoSlopeNorm(vmin=vmin, vcenter=0, vmax=max(vmax, 0.1))
    cmap2 = LinearSegmentedColormap.from_list("rb", ["#1d4ed8", "#0b1220", "#ef4444"])
    im = ax.imshow(rt.values, aspect="auto", cmap=cmap2, norm=norm_rb)
    ax.set_xticks(range(len(list(rt.columns))))
    ax.set_xticklabels(list(rt.columns), color=TEXT)
    ax.set_yticks(range(len(list(rt.index))))
    ax.set_yticklabels(list(rt.index), color=TEXT)
    for i in range(rt.shape[0]):
        for j in range(rt.shape[1]):
            ax.text(j, i, f"{rt.values[i, j]:.1f}", ha="center", va="center", color="white", fontsize=7)
    cb = fig.colorbar(im, ax=ax, fraction=0.03)
    cb.ax.yaxis.set_tick_params(color=TEXT)
    plt.setp(cb.ax.yaxis.get_ticklabels(), color=TEXT)
    ax.set_title("下單時段 × 客戶屬性 標準化殘差（紅＝偏多／藍＝偏少）", color=CYAN, fontsize=12)
    resid_time_b64 = fig_b64(fig)

    labels_r, lifts = [], []
    for r in rules:
        labels_r.append("、".join(r["antecedents"]) + " → " + "、".join(r["consequents"]))
        lifts.append(r["lift"])
    if labels_r:
        order_idx = np.argsort(lifts)[::-1]
        labels_s = [labels_r[i] for i in order_idx]
        lifts_s = [lifts[i] for i in order_idx]
    else:
        labels_s, lifts_s = ["（無標黃規則）"], [0]
    fig, ax = plt.subplots(figsize=(10, 5.2), facecolor=BG)
    ax.set_facecolor(BG)
    y = np.arange(len(labels_s))
    ax.barh(y, lifts_s, color=CYAN, height=0.55)
    ax.set_yticks(y)
    ax.set_yticklabels(labels_s, color=TEXT, fontsize=9)
    ax.invert_yaxis()
    ax.tick_params(colors=TEXT)
    ax.set_xlabel("Lift（搭售提升倍數）", color=MUTED)
    for yi, v in zip(y, lifts_s):
        ax.text(v + 0.3, yi, f"{v:.2f} 倍", va="center", color=CYAN, fontsize=8)
    ax.set_title("搭售規則 · 售提升倍數對比", color=CYAN, fontsize=12)
    for spine in ax.spines.values():
        spine.set_color("#334")
    ax.grid(axis="x", color="#233", alpha=0.5)
    lift_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(11, 4.2), facecolor=BG)
    ax.set_facecolor(BG)
    if len(cust):
        cplot = cust.sort_values(["orders", "amount"], ascending=[False, False])
        x = np.arange(len(cplot))
        ax.bar(x, cplot["orders"], color=PINK, width=0.55, label="下單次數")
        ax2 = ax.twinx()
        ax2.plot(x, cplot["amount"] / 10000, color=YELLOW, marker="o", linewidth=2, label="銷售額(萬港元)")
        ax.set_xticks(x)
        ax.set_xticklabels(cplot["customer"], rotation=45, ha="right", color=TEXT, fontsize=8)
        ax2.tick_params(colors=TEXT)
        ax2.set_ylabel("銷售額（萬港元）", color=YELLOW)
        for spine in list(ax2.spines.values()):
            spine.set_color("#334")
    else:
        ax.text(0.5, 0.5, "本週無客戶下單", ha="center", va="center", color=MUTED, transform=ax.transAxes)
    ax.tick_params(colors=TEXT)
    ax.set_ylabel("下單次數", color=PINK)
    ax.set_title("客戶下單次數與銷售額", color=CYAN, fontsize=12)
    for spine in ax.spines.values():
        spine.set_color("#334")
    cust_chart_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(11, 4.2), facecolor=BG)
    ax.set_facecolor(BG)
    if len(prod):
        pplot = prod.head(16)
        x = np.arange(len(pplot))
        ax.bar(x, pplot["boxes"], color=PINK, width=0.55)
        ax2 = ax.twinx()
        ax2.plot(x, pplot["orders"], color=YELLOW, marker="o", linewidth=2)
        ax.set_xticks(x)
        ax.set_xticklabels(pplot["product"], rotation=45, ha="right", color=TEXT, fontsize=8)
        ax2.tick_params(colors=TEXT)
        ax2.set_ylabel("品項下單次數", color=YELLOW)
        for spine in list(ax2.spines.values()):
            spine.set_color("#334")
    else:
        ax.text(0.5, 0.5, "本週無產品下單", ha="center", va="center", color=MUTED, transform=ax.transAxes)
    ax.tick_params(colors=TEXT)
    ax.set_ylabel("下單箱數", color=PINK)
    ax.set_title("產品下單統計", color=CYAN, fontsize=12)
    for spine in ax.spines.values():
        spine.set_color("#334")
    prod_chart_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(11, 3.8), facecolor=BG)
    ax.set_facecolor(BG)
    cmap_h = LinearSegmentedColormap.from_list("h", ["#fff7ed", "#fdba74", "#ea580c", "#7f1d1d"])
    im = ax.imshow(heat.values, aspect="auto", cmap=cmap_h)
    ax.set_xticks(range(15))
    ax.set_xticklabels(list(range(6, 21)), color=TEXT)
    ax.set_yticks(range(6))
    ax.set_yticklabels(["星期一", "星期二", "星期三", "星期四", "星期五", "星期六"], color=TEXT)
    for i in range(6):
        for j in range(15):
            v = int(heat.values[i, j])
            if v:
                ax.text(
                    j,
                    i,
                    str(v),
                    ha="center",
                    va="center",
                    color="#111" if v < heat.values.max() * 0.55 else "white",
                    fontsize=7,
                )
    cb = fig.colorbar(im, ax=ax, fraction=0.03)
    cb.set_label("訂單量", color=TEXT)
    cb.ax.yaxis.set_tick_params(color=TEXT)
    plt.setp(cb.ax.yaxis.get_ticklabels(), color=TEXT)
    ax.set_title("訂單量熱力分佈（週一至週六 × 06:00–20:00｜歷史全量｜同一時間計一次）", color=CYAN, fontsize=11)
    heat_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(6.2, 4.8), facecolor=BG)
    ax.set_facecolor(BG)
    colors = ["#22d3ee", "#a78bfa", "#f472b6", "#fbbf24", "#334155"]
    wedges, _, autotexts = ax.pie(
        pareto_vals,
        labels=None,
        colors=colors[: len(pareto_vals)],
        autopct=lambda p: f"{p:.1f}%" if p > 3 else "",
        pctdistance=0.75,
        wedgeprops=dict(width=0.42, edgecolor=BG, linewidth=2),
    )
    for t in autotexts:
        t.set_color("white")
        t.set_fontsize(9)
    ax.legend(
        wedges,
        [f"{l} {v * 100:.1f}%" for l, v in zip(pareto_labels, pareto_vals)],
        loc="center left",
        bbox_to_anchor=(1.0, 0.5),
        facecolor=CARD,
        edgecolor="#334",
        labelcolor=TEXT,
        fontsize=9,
    )
    ax.set_title("頭部客戶金額貢獻佔比", color=CYAN, fontsize=12)
    pareto_b64 = fig_b64(fig)

    fig, ax = plt.subplots(figsize=(10, 4.2), facecolor=BG)
    ax.set_facecolor(BG)
    x = np.arange(len(bin_labels))
    ax.bar(x, hist_counts.values, color="#fb7185", width=0.65, label="訂單筆數")
    ax.plot(x, kde_scaled, color=RED, marker="o", linewidth=2, label="KDE")
    ax.set_xticks(x)
    ax.set_xticklabels(bin_labels, color=TEXT, fontsize=9)
    ax.tick_params(colors=TEXT)
    ax.set_ylabel("訂單筆數", color=TEXT)
    ax.set_title("客戶下單金額分布（直方圖 + 核密度估計）", color=CYAN, fontsize=12)
    ax.legend(facecolor=CARD, edgecolor="#334", labelcolor=TEXT)
    for spine in ax.spines.values():
        spine.set_color("#334")
    ax.grid(axis="y", color="#233", alpha=0.5)
    amt_b64 = fig_b64(fig)

    # insights
    flat = [(resid_plot.loc[a, c], a, c) for a in resid_plot.index for c in resid_plot.columns]
    flat.sort(reverse=True)
    insight_bits = [f"「{a}」對「{c}」偏好偏高（殘差 {v:.1f}）" for v, a, c in flat[:6] if v >= 2]
    resid_insight = "；".join(insight_bits[:4]) if insight_bits else "本週殘差結構整體平穩"

    tflat = [(rt.loc[a, h], a, h) for a in rt.index for h in rt.columns]
    tflat.sort(reverse=True)
    time_parts = [f"「{a}」於 {h}:00 偏多（殘差 {v:.1f}）" for v, a, h in tflat[:5] if v >= 2]
    time_insight = "；".join(time_parts[:4]) if time_parts else "下單時段集中於上午"

    def rule_card(r: dict) -> str:
        ant = "、".join(r["antecedents"])
        con = "、".join(r["consequents"])
        extra = ""
        if r.get("common_category") is not None:
            extra += f'<div class="extra">common_category：{r["common_category"]}（同中類）</div>'
        if r.get("category_internal_support") is not None:
            extra += (
                f'<div class="extra">category_internal_support：'
                f'{r["category_internal_support"] * 100:.2f}%（同中類訂單內同時購買佔比）</div>'
            )
        return f"""<div class="rule-card panel">
      <div class="rule-title">{ant} → {con}</div>
      <div class="metric">組合佔比（Support）：{r["support"] * 100:.2f}%（同單同時購買佔比）</div>
      <div class="metric">搭購轉化率（Confidence）：{r["confidence"] * 100:.1f}%（買了前者亦買後者）</div>
      <div class="metric">搭售提升倍數（Lift）：{r["lift"]:.2f} 倍（相對平常購買機率）</div>
      {extra}
    </div>"""

    attr_colors = {"飛機": "#38bdf8", "加工": "#a78bfa", "凍肉店": "#fb7185", "燒臘": "#fbbf24", "肉檯": "#4ade80"}
    attr_cards = ""
    if n_orders:
        for a, c in attr_counts.items():
            share = c / n_orders
            attr_cards += f"""<div class="attr-card panel"><div class="attr-name" style="color:{attr_colors.get(a, CYAN)}">{a}</div>
      <div class="attr-pct">{share * 100:.1f}%</div>
      <div class="attr-sub">{c}/{n_orders} 單</div></div>"""
    else:
        attr_cards = '<div class="panel note">本週無已登記訂單，暫無客戶屬性佔比。</div>'

    prod_detail = "、".join(
        [
            f"{r['product']}（{int(r['boxes']) if float(r['boxes']).is_integer() else r['boxes']} / {int(r['orders'])}）"
            for _, r in prod.iterrows()
        ]
    ) or "（無）"

    group_html = ""
    for i, g in enumerate(groups, 1):
        names = "、".join(g["customers"]) if g["customers"] else "（無）"
        m = g["monetary"]
        m_txt = f"{m / 10000:.1f} 萬" if m >= 10000 else f"{m:,.2f}"
        group_html += f"""<div class="k-card panel">
      <div class="k-title">第{i}組 · {g['name']}</div>
      <div class="k-metric">距今最近一次交易天數：<b>{g['recency']:.2f}</b></div>
      <div class="k-metric">交易頻次：<b>{g['frequency']:.2f}</b></div>
      <div class="k-metric">交易金額：<b>{m_txt}</b></div>
      <div class="k-list">客戶明細：{names}</div>
    </div>"""

    amt_table_rows = "".join(
        f"<tr><td>{lab}</td><td>{int(c)}</td><td>{k:.3f}</td></tr>"
        for lab, c, k in zip(bin_labels, hist_counts.values, kde_scaled)
    )
    new_cust_txt = "、".join(new_customers) if new_customers else "（無）"
    if prev["customers"]:
        cust_delta = (
            "持平"
            if n_customers == prev["customers"]
            else f'變化 {(n_customers - prev["customers"]) / prev["customers"] * 100:+.1f}%'
        )
    else:
        cust_delta = "—"

    resid_cat_json = json.dumps(
        {
            "z": resid_full.values.astype(float).tolist(),
            "x": [str(c) for c in resid_full.columns.tolist()],
            "y": [str(i) for i in resid_full.index.tolist()],
        },
        ensure_ascii=False,
    )

    amount_disp = f"{total_amount / 10000:.1f}萬" if total_amount >= 1000 else f"{total_amount:,.2f}"
    boxes_disp = int(total_boxes) if float(total_boxes).is_integer() else total_boxes
    if rules_mode.startswith("product_level_rules"):
        rules_source = rules_path.name
    else:
        rules_source = f"{rules_path.name}（product_level_rules 工作表）"

    html = f"""<!DOCTYPE html>
<html lang="zh-Hant">
<head>
<meta charset="UTF-8" />
<meta name="viewport" content="width=device-width, initial-scale=1" />
<title>2026年第29周(7.13-7.18) B端商城運營分析報告</title>
<script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/html2canvas@1.4.1/dist/html2canvas.min.js"></script>
<style>
:root {{
  --bg:#0a0e14; --panel:#101820; --line:#1f3a44; --cyan:#2ee6d6; --pink:#ff4f8b;
  --yellow:#ffd166; --text:#e8eef7; --muted:#9aa8bc;
  --body-scale:1; --head-scale:1;
}}
* {{ box-sizing:border-box; }}
body {{
  margin:0; font-family:"Microsoft JhengHei","PingFang TC","Noto Sans TC","WenQuanYi Micro Hei",sans-serif;
  background:radial-gradient(1200px 600px at 10% -10%, #122033 0%, var(--bg) 55%), var(--bg);
  color:var(--text); line-height:1.55; font-size:calc(15px * var(--body-scale));
}}
.wrap {{ max-width:1120px; margin:0 auto; padding:28px 18px 120px; }}
h1 {{
  color:var(--cyan); font-size:calc(30px * var(--head-scale)); text-align:center;
  margin:0 0 6px; letter-spacing:1px;
}}
.sub {{ color:var(--muted); font-size:13px; margin-bottom:10px; text-align:center; }}
.nav {{
  display:flex; flex-wrap:wrap; gap:8px; justify-content:center; margin:12px 0 22px;
}}
.nav a {{
  color:var(--cyan); text-decoration:none; border:1px solid #1f4a52;
  padding:4px 10px; border-radius:999px; font-size:12px; background:rgba(10,30,40,.55);
}}
h2.sec {{
  color:var(--cyan); font-size:calc(22px * var(--head-scale));
  margin:34px 0 10px; border-bottom:1px solid var(--line); padding-bottom:8px;
}}
.grid {{ display:grid; gap:12px; }}
.kpi-grid {{ grid-template-columns:repeat(4,1fr); }}
.panel {{
  background:linear-gradient(180deg, #101a27, #0b121c);
  border:1px solid var(--line); border-radius:14px; padding:14px 16px;
}}
.panel.light {{ background:#eef1f5; color:#111; }}
.kpi .label {{ color:var(--muted); font-size:12px; }}
.kpi .value {{ font-size:26px; font-weight:700; color:#fff; margin-top:4px; }}
.kpi .unit {{ font-size:12px; color:var(--muted); }}
.wow {{ font-size:12px; margin-top:6px; }}
.wow.up {{ color:#4ade80; }}
.wow.down {{ color:#fb7185; }}
.note {{ color:var(--muted); font-size:13px; margin:10px 0 0; }}
.warn {{ color:#fbbf24; font-size:13px; margin:8px 0 0; }}
.editable {{
  border:1px dashed #2f6a74; background:#0a1620; border-radius:8px;
  padding:6px 8px; min-height:28px; outline:none;
}}
.editable:focus {{ border-color:var(--cyan); }}
.attr-grid {{ grid-template-columns:repeat(3,1fr); }}
.attr-card {{ text-align:center; padding:18px 10px; }}
.attr-name {{ font-size:18px; font-weight:700; }}
.attr-pct {{ font-size:32px; font-weight:800; margin:6px 0; }}
.attr-sub {{ color:var(--muted); font-size:13px; }}
.img {{ width:100%; border-radius:12px; border:1px solid var(--line); background:#05080e; }}
.rules {{ grid-template-columns:1fr 1fr; }}
.rule-title {{ font-weight:700; margin-bottom:8px; }}
.metric {{ color:var(--cyan); font-size:13px; margin:3px 0; }}
.extra {{ color:#86efac; font-size:12px; margin-top:4px; }}
.k-grid {{ grid-template-columns:repeat(4,1fr); }}
.k-card {{ min-height:220px; }}
.k-title {{ color:var(--cyan); font-weight:700; margin-bottom:8px; }}
.k-metric {{ font-size:13px; margin:4px 0; }}
.k-list {{ margin-top:8px; color:var(--muted); font-size:12px; word-break:break-all; }}
table {{ width:100%; border-collapse:collapse; font-size:13px; }}
th,td {{ border:1px solid var(--line); padding:8px 10px; text-align:left; }}
th {{ background:#121c2a; color:var(--cyan); }}
.summary-block {{ margin:10px 0 16px; }}
.summary-block h3 {{ margin:0 0 6px; color:#fff; font-size:calc(16px * var(--head-scale)); }}
.trad .value {{ font-size:18px; }}
#residCatPlot {{ width:100%; min-height:420px; }}
.resid-tools {{
  display:flex; gap:16px; align-items:center; flex-wrap:wrap;
  margin:8px 0 4px; color:#c9d4e3; font-size:12px;
}}
.resid-tools input[type=range] {{ width:180px; }}
.toolbar {{
  position:fixed; left:0; right:0; bottom:0; z-index:50;
  background:rgba(6,12,20,.92); border-top:1px solid #1f3a44;
  backdrop-filter:blur(8px); padding:10px 14px 12px;
}}
.toolbar-inner {{ max-width:1120px; margin:0 auto; }}
.font-box {{
  border:1px solid #244; border-radius:10px; padding:8px 10px;
  margin-bottom:8px; background:rgba(10,20,32,.8);
}}
.font-row {{
  display:flex; align-items:center; gap:8px; flex-wrap:wrap; margin:4px 0; font-size:13px;
}}
.font-row input[type=range] {{ width:160px; }}
.btns {{ display:flex; flex-wrap:wrap; gap:8px; }}
.btn {{
  border-radius:10px; padding:7px 12px; font-size:13px; cursor:pointer;
  background:rgba(8,16,24,.75);
}}
.btn.yellow {{ border:1px solid #fbbf24; color:#fbbf24; }}
.btn.green {{ border:1px solid #34d399; color:#34d399; }}
.btn.cyan {{ border:1px solid #22d3ee; color:#22d3ee; }}
.btn.pink {{ border:1px solid #fb7185; color:#fb7185; }}
.btn:hover {{ filter:brightness(1.15); }}
.hint {{ color:#8b9bb4; font-size:11px; margin:6px 0 8px; }}
.pdf-mode .toolbar {{ display:none !important; }}
@media (max-width:900px) {{
  .kpi-grid,.attr-grid,.rules,.k-grid {{ grid-template-columns:1fr 1fr; }}
}}
@media (max-width:600px) {{
  .kpi-grid,.attr-grid,.rules,.k-grid {{ grid-template-columns:1fr; }}
}}
@media print {{
  .toolbar {{ display:none !important; }}
  body {{ background:#fff !important; color:#111 !important; }}
  h1,h2.sec,.nav a,.metric,.k-title {{ color:#0b7285 !important; }}
  .kpi,.panel,.rule-card,.k-card,.attr-card,.summary-block {{ break-inside:avoid; }}
}}
</style>
</head>
<body>
<div class="wrap" id="reportRoot">
  <h1>2026年第29周(7.13-7.18) B端商城運營分析報告</h1>
  <div class="sub">推送效果 · 客戶轉化 · 產品下單 · 關聯規則 · 殘差與聚類 · 下週計劃　｜　v1 · week29_ops_report.html</div>
  <div class="nav">
    <a href="#s1">推送成效</a><a href="#s2">客戶轉化</a><a href="#s3">雷達圖</a><a href="#s4">客戶屬性</a>
    <a href="#s5">殘差分析</a><a href="#s6">K-means</a><a href="#s7">關聯規則</a><a href="#s8">客戶/產品</a>
    <a href="#s10">熱力/二八</a><a href="#s12">金額分布</a><a href="#s13">整體情況</a>
  </div>
  {"<p class='warn'>" + data_note + "</p>" if data_note else ""}

  <h2 class="sec" id="s1">一、核心指標總覽</h2>
  <div class="grid kpi-grid">
    <div class="panel kpi"><div class="label">目標完成率</div><div class="value">{target_rate * 100:.2f}%</div><div class="unit">累計 {accum / 1e4:.1f} 萬 / 目標 {target / 1e4:.0f} 萬</div></div>
    <div class="panel kpi"><div class="label">推送總次數</div><div class="value">{push}</div>{wow_html(push, prev["push"])}</div>
    <div class="panel kpi"><div class="label">點擊總次數</div><div class="value">{clicks}</div>{wow_html(clicks, prev["clicks"])}</div>
    <div class="panel kpi"><div class="label">成功接收設備數</div><div class="value">{devices}</div>{wow_html(devices, prev["devices"])}</div>
    <div class="panel kpi"><div class="label">APP 下單次數</div><div class="value">{n_orders}</div>{wow_html(n_orders, prev["orders"])}<div class="unit">佔總下單次數比例（可編輯）</div><div class="editable" contenteditable="true" data-key="ratio_orders">請手動填寫</div></div>
    <div class="panel kpi"><div class="label">總金額（港元）</div><div class="value">{amount_disp}</div>{wow_html(total_amount, prev["amount"])}<div class="unit">佔總下單金額（可編輯）</div><div class="editable" contenteditable="true" data-key="ratio_amount">請手動填寫</div></div>
    <div class="panel kpi trad"><div class="label">傳統渠道 APP 佔比（可編輯）</div><div class="value">金額佔比 / 次數佔比</div>
      <div style="display:flex;gap:8px;flex-wrap:wrap;margin-top:6px;">
        <div>金額：<span class="editable" contenteditable="true" data-key="trad_amt" style="display:inline-block;min-width:48px;">____%</span></div>
        <div>次數：<span class="editable" contenteditable="true" data-key="trad_cnt" style="display:inline-block;min-width:48px;">____%</span></div>
      </div>
    </div>
    <div class="panel kpi"><div class="label">下單箱數 / 下單客戶</div><div class="value">{boxes_disp} / {n_customers}</div>{wow_html(n_customers, prev["customers"])}</div>
  </div>
  <p class="note">本週篩選：R列時間 ∈ 2026-07-13～2026-07-18、P列＝已登記、O列非空；APP下單次數與總金額按 ERP 單號去重。產品名稱取「小類」（對應原報告 T/U 產品維度；本檔因多出創建人欄位後為 Y 列）、箱數取 I列。客戶屬性取自工作簿「客戶屬性1」對照（本檔 VLOOKUP 結果異常，已自動校正）。本週共 {n_orders} 單、{boxes_disp} 箱、{n_customers} 家客戶。數據檔：{order_path.name}。</p>

  <h2 class="sec" id="s2">二、轉化流程網狀圖</h2>
  <img class="img" src="data:image/png;base64,{funnel_b64}" alt="轉化流程" />

  <h2 class="sec" id="s3">三、多維度雷達圖</h2>
  <img class="img" src="data:image/png;base64,{radar_b64}" alt="雷達圖" />

  <h2 class="sec" id="s4">四、客戶屬性下單佔比</h2>
  <div class="grid attr-grid">{attr_cards}</div>
  <p class="note">取數：客戶屬性（客戶屬性1 對照／校正後）；O列 ERP 單號去重後計次。</p>

  <h2 class="sec" id="s5">五、殘差分析</h2>
  <div class="sub" style="text-align:left">小類 × 客戶屬性（調整 |殘差| 範圍時，不在範圍內的產品直接隱藏）</div>
  <div class="resid-tools">
    <span>|殘差| 下限 <b id="rminLab">2.0</b></span>
    <input id="rmin" type="range" min="0" max="10" step="0.1" value="2.0" />
    <span>|殘差| 上限 <b id="rmaxLab">16.0</b></span>
    <input id="rmax" type="range" min="2" max="20" step="0.1" value="16.0" />
  </div>
  <div class="panel light"><div id="residCatPlot"></div></div>
  <p class="note">取數來源：{frozen_path.name}（客戶屬性 × 小类）。調整 |殘差| 範圍後，不在範圍內的產品（小類）會直接隱藏。{resid_insight}。</p>
  <div class="sub" style="text-align:left;margin-top:18px">下單時段 × 客戶屬性</div>
  <div class="panel light"><img class="img" src="data:image/png;base64,{resid_time_b64}" alt="時段殘差" /></div>
  <p class="note">取數來源：{order_path.name}（歷史已登記訂單，O列去重）。{time_insight}。</p>

  <h2 class="sec" id="s6">六、K-means 客戶聚類分析</h2>
  <div class="grid k-grid">{group_html}</div>
  <p class="note">特徵：距今最近一次交易天數（R）、交易頻次（O唯一）、交易金額（L，同ERP不重複）；基於累計已登記訂單做 4 組聚類。</p>

  <h2 class="sec" id="s7">七、關聯規則 · 重點品項搭售組合</h2>
  <p class="note">本週依 {rules_source} 中標黃組合更新；若同中類子表有對應指標，則附加 common_category 與 category_internal_support。<br/>
  Support＝組合佔比；Confidence＝搭購轉化率；Lift＝相對平常購買機率的提升倍數（&gt;1 為正向關聯）。</p>
  <div class="grid rules">{"".join(rule_card(r) for r in rules) if rules else '<div class="panel note">未偵測到標黃組合。</div>'}</div>
  <img class="img" style="margin-top:12px" src="data:image/png;base64,{lift_b64}" alt="Lift對比" />

  <h2 class="sec" id="s8">八、客戶下單次數與銷售額</h2>
  <p class="note">按客戶名稱（E列）彙總，共 {n_customers} 家；O列去重後 {n_orders} 單，總金額 {total_amount:,.2f} 港元（較上週客戶數 {prev["customers"]} 家 {cust_delta}）。</p>
  <div class="panel" style="margin-bottom:12px"><b>本週新增下單用戶（共 {len(new_customers)} 家）：</b>{new_cust_txt}<div class="note" style="margin-top:6px">自動比對歷史已登記客戶名單；僅本週首次出現者列入。</div></div>
  <img class="img" src="data:image/png;base64,{cust_chart_b64}" alt="客戶下單" />

  <h2 class="sec" id="s9">九、產品下單統計</h2>
  <p class="note">產品名稱取「小類」，箱數加總 I列；本週合計 {boxes_disp} 箱。</p>
  <img class="img" src="data:image/png;base64,{prod_chart_b64}" alt="產品統計" />
  <p class="note">明細：{prod_detail}</p>

  <h2 class="sec" id="s10">十、圖一．訂單量熱力分佈（週一至週六 × 06:00–20:00）</h2>
  <p class="note">使用全部歷史已登記數據；R列時間，同一時間只計一次；細分至星期 × 小時（06:00–20:00）。</p>
  <img class="img" src="data:image/png;base64,{heat_b64}" alt="熱力圖" />

  <h2 class="sec" id="s11">十一、圖二．客戶金額貢獻（頭部 20% 客戶／二八）</h2>
  <p class="note">本週按銷售額排序，頭部約 20% 客戶（{n_top} 家）貢獻總金額的 <b>{top_share * 100:.1f}%</b>；金額取 L列、客戶取 E列，同 ERP 不重複。</p>
  <img class="img" src="data:image/png;base64,{pareto_b64}" alt="二八" />
  <div class="panel" style="margin-top:12px">
    <div style="color:var(--cyan);font-weight:700;margin-bottom:6px;">結論（可編輯）</div>
    <div class="editable" contenteditable="true" data-key="pareto_note">請在此書寫本週二八集中度結論……</div>
  </div>

  <h2 class="sec" id="s12">十二、圖五．客戶下單金額分布（直方圖＋核密度估計）</h2>
  <p class="note">歷史全量按 O列去重共 {len(amt_hist)} 筆訂單、{amt_hist["amount"].nunique()} 個不同金額（L列）；長條為頻次，紅線為 KDE。</p>
  <img class="img" src="data:image/png;base64,{amt_b64}" alt="金額分布" />
  <table style="margin-top:12px">
    <thead><tr><th>金額區間（港元）</th><th>訂單筆數</th><th>KDE</th></tr></thead>
    <tbody>{amt_table_rows}</tbody>
  </table>

  <h2 class="sec" id="s13">十三、本週整體情況分析</h2>
  <div class="summary-block panel">
    <h3>運營情況</h3>
    <div class="editable" contenteditable="true" data-key="ops">本週推送 {push} 次、成功接收設備 {devices} 個、點擊 {clicks} 次；APP 去重下單 {n_orders} 次。請補充運營側觀察與問題跟進……</div>
  </div>
  <div class="summary-block panel">
    <h3>平台銷售額</h3>
    <div class="editable" contenteditable="true" data-key="sales">本週 APP 下單金額 {total_amount:,.2f} 港元（約 {total_amount / 10000:.2f} 萬），較上週 {prev["amount"] / 10000:.1f} 萬 {(total_amount - prev["amount"]) / 10000:+.2f} 萬。頭部 20% 客戶貢獻 {top_share * 100:.1f}%。請補充銷售解讀……</div>
  </div>
  <div class="summary-block panel">
    <h3>下週計劃</h3>
    <div class="editable" contenteditable="true" data-key="plan">請填寫下週促銷節奏、重點客群與品項搭售實驗計劃……</div>
  </div>
</div>

<div class="toolbar" id="toolbar">
  <div class="toolbar-inner">
    <div class="font-box">
      <div class="font-row">正文字體
        <button class="btn cyan" type="button" onclick="nudgeFont('body',-0.05)">A-</button>
        <input id="bodyRange" type="range" min="0.8" max="1.4" step="0.01" value="1" oninput="setFont('body', this.value)" />
        <button class="btn cyan" type="button" onclick="nudgeFont('body',0.05)">A+</button>
        <span id="bodyPct">100%</span>
      </div>
      <div class="font-row">標題字體
        <button class="btn cyan" type="button" onclick="nudgeFont('head',-0.05)">A-</button>
        <input id="headRange" type="range" min="0.8" max="1.4" step="0.01" value="1" oninput="setFont('head', this.value)" />
        <button class="btn cyan" type="button" onclick="nudgeFont('head',0.05)">A+</button>
        <span id="headPct">100%</span>
        <button class="btn green" type="button" onclick="resetFont()">全部重設 100%</button>
      </div>
    </div>
    <div class="hint">若自動 PDF 匯出效果不理想，可用「截圖助手」+「打開空白組裝頁」手動貼上截圖做緊湊排版。</div>
    <div class="btns">
      <button class="btn yellow" type="button" onclick="screenshotAssist()">◆ 截圖助手</button>
      <button class="btn yellow" type="button" onclick="openAssemblePage()">◆ 打開空白組裝頁</button>
      <button class="btn green" type="button" onclick="togglePdfMode()">◆ PDF 排版模式</button>
      <button class="btn green" type="button" onclick="exportBlocksPdf()">◆ 按區塊匯出 PDF</button>
      <button class="btn cyan" type="button" onclick="exportPdf()">◆ 匯出為 PDF</button>
      <button class="btn cyan" type="button" onclick="window.print()">◆ 列印 / 另存 PDF</button>
      <button class="btn pink" type="button" onclick="clearStorage()">◆ 清除暫存</button>
    </div>
  </div>
</div>

<script>
const RESID_CAT = {resid_cat_json};
const STORAGE_KEY = 'week29_ops_report_v1';

function setFont(kind, val) {{
  val = Math.min(1.4, Math.max(0.8, parseFloat(val)));
  if (kind === 'body') {{
    document.documentElement.style.setProperty('--body-scale', val);
    document.getElementById('bodyRange').value = val;
    document.getElementById('bodyPct').textContent = Math.round(val*100) + '%';
  }} else {{
    document.documentElement.style.setProperty('--head-scale', val);
    document.getElementById('headRange').value = val;
    document.getElementById('headPct').textContent = Math.round(val*100) + '%';
  }}
  persist();
}}
function nudgeFont(kind, d) {{
  const el = document.getElementById(kind === 'body' ? 'bodyRange' : 'headRange');
  setFont(kind, parseFloat(el.value) + d);
}}
function resetFont() {{ setFont('body', 1); setFont('head', 1); }}
function persist() {{
  const data = {{ body: document.getElementById('bodyRange').value, head: document.getElementById('headRange').value, edits: {{}} }};
  document.querySelectorAll('[data-key]').forEach(el => {{ data.edits[el.dataset.key] = el.innerHTML; }});
  localStorage.setItem(STORAGE_KEY, JSON.stringify(data));
}}
function restore() {{
  try {{
    const data = JSON.parse(localStorage.getItem(STORAGE_KEY) || '{{}}');
    if (data.body) setFont('body', data.body);
    if (data.head) setFont('head', data.head);
    if (data.edits) {{
      Object.entries(data.edits).forEach(([k,v]) => {{
        const el = document.querySelector('[data-key="'+k+'"]');
        if (el) el.innerHTML = v;
      }});
    }}
  }} catch(e) {{}}
}}
function clearStorage() {{
  localStorage.removeItem(STORAGE_KEY);
  alert('已清除本頁暫存（字級與可編輯文字）。重新整理後恢復預設。');
}}
document.querySelectorAll('[contenteditable]').forEach(el => el.addEventListener('input', persist));

function renderResidCat() {{
  const rmin = parseFloat(document.getElementById('rmin').value);
  const rmax = parseFloat(document.getElementById('rmax').value);
  document.getElementById('rminLab').textContent = rmin.toFixed(1);
  document.getElementById('rmaxLab').textContent = rmax.toFixed(1);
  const filtered = RESID_CAT.z.map(row => row.map(v => {{
    const a = Math.abs(v);
    return (a >= rmin && a <= rmax) ? v : null;
  }}));
  const keepCols = [];
  for (let j = 0; j < RESID_CAT.x.length; j++) {{
    const hasInRange = filtered.some(row => row[j] != null && Number.isFinite(row[j]));
    if (hasInRange) keepCols.push(j);
  }}
  const x = keepCols.map(j => RESID_CAT.x[j]);
  const z = filtered.map(row => keepCols.map(j => row[j]));
  const y = RESID_CAT.y;
  const heat = {{
    z, x, y, type:'heatmap',
    colorscale:[[0,'#1d4ed8'],[0.5,'#f8fafc'],[1,'#dc2626']],
    zmid:0, hoverongaps:false, colorbar:{{title:'殘差'}}
  }};
  Plotly.newPlot('residCatPlot', [heat], {{
    margin:{{t:20,l:70,r:30,b:80}},
    paper_bgcolor:'#eef1f5', plot_bgcolor:'#eef1f5', font:{{color:'#111'}},
    xaxis:{{tickangle:-45, automargin:true}}, yaxis:{{automargin:true}},
    annotations: keepCols.length ? [] : [{{
      text: '當前殘差範圍內無產品可顯示', xref:'paper', yref:'paper',
      x:0.5, y:0.5, showarrow:false, font:{{size:14, color:'#666'}}
    }}]
  }}, {{responsive:true, displayModeBar:true}});
}}
document.getElementById('rmin').addEventListener('input', renderResidCat);
document.getElementById('rmax').addEventListener('input', renderResidCat);

async function screenshotAssist() {{
  const root = document.getElementById('reportRoot');
  const canvas = await html2canvas(root, {{ backgroundColor:'#0a0e14', scale:2, useCORS:true }});
  canvas.toBlob(blob => {{
    const url = URL.createObjectURL(blob);
    const a = document.createElement('a');
    a.href = url; a.download = 'week29_report_screenshot.png'; a.click();
    URL.revokeObjectURL(url);
  }});
}}
function openAssemblePage() {{
  const w = window.open('', '_blank');
  w.document.write(`<!DOCTYPE html><html><head><meta charset="UTF-8"><title>空白組裝頁</title>
    <style>body{{margin:0;background:#111;color:#eee;font-family:sans-serif}}
    .box{{max-width:900px;margin:20px auto;padding:16px}}
    .drop{{min-height:70vh;border:2px dashed #2ee6d6;border-radius:12px;padding:16px}}
    img{{max-width:100%;display:block;margin:8px 0}}</style></head><body>
    <div class="box"><h2>空白組裝頁</h2><p>將截圖拖曳/貼上到下方區域，再列印另存 PDF。</p>
    <div class="drop" id="drop" contenteditable="true"></div></div>
    <script>const d=document.getElementById('drop');
    d.addEventListener('paste',e=>{{const items=e.clipboardData&&e.clipboardData.items; if(!items)return;
      for(const it of items){{ if(it.type.startsWith('image/')){{ const f=it.getAsFile(); const r=new FileReader();
        r.onload=ev=>{{ const img=document.createElement('img'); img.src=ev.target.result; d.appendChild(img); }}; r.readAsDataURL(f); e.preventDefault(); }} }} }});
    <\\/script></body></html>`);
  w.document.close();
}}
function togglePdfMode() {{
  document.documentElement.classList.toggle('pdf-mode');
  alert(document.documentElement.classList.contains('pdf-mode')
    ? '已進入 PDF 排版模式（隱藏底欄）。再次點擊可退出。'
    : '已退出 PDF 排版模式。');
}}
function exportPdf() {{
  document.documentElement.classList.add('pdf-mode');
  setTimeout(() => window.print(), 200);
}}
async function exportBlocksPdf() {{
  const secs = [...document.querySelectorAll('h2.sec')];
  for (let i=0;i<secs.length;i++) {{
    const start = secs[i];
    const end = secs[i+1] || null;
    const range = document.createRange();
    range.setStartBefore(start);
    if (end) range.setEndBefore(end);
    else range.setEndAfter(document.getElementById('reportRoot').lastElementChild);
    const div = document.createElement('div');
    div.appendChild(range.cloneContents());
    div.style.background='#0a0e14'; div.style.padding='16px'; div.style.width='1100px';
    document.body.appendChild(div);
    const canvas = await html2canvas(div, {{ backgroundColor:'#0a0e14', scale:1.5 }});
    document.body.removeChild(div);
    const a = document.createElement('a');
    a.href = canvas.toDataURL('image/png');
    a.download = `week29_block_${{String(i+1).padStart(2,'0')}}.png`;
    a.click();
  }}
  alert('已按區塊匯出 PNG（可再組裝為 PDF）。');
}}
restore();
renderResidCat();
</script>
</body>
</html>
"""

    out = REPORT_DIR / "2026_week29_ops_report.html"
    out.write_text(html, encoding="utf-8")
    # also copy to artifacts if available
    art = Path("/opt/cursor/artifacts")
    if art.exists():
        (art / "2026_week29_ops_report.html").write_text(html, encoding="utf-8")

    summary = {
        "order_file": order_path.name,
        "frozen_file": frozen_path.name,
        "rules_source": rules_source,
        "orders": n_orders,
        "amount": round(total_amount, 2),
        "customers": n_customers,
        "boxes": total_boxes,
        "new_customers": new_customers,
        "attr": attr_counts.to_dict() if n_orders else {},
        "rules": len(rules),
        "rules_with_common_category": sum(1 for r in rules if r.get("common_category") is not None),
        "rules_with_cis": sum(1 for r in rules if r.get("category_internal_support") is not None),
        "groups": [
            (
                g["name"],
                len(g["customers"]),
                round(g["recency"], 2),
                round(g["frequency"], 2),
                round(g["monetary"], 2),
            )
            for g in groups
        ],
        "pareto_share": round(top_share, 4),
        "target_rate": round(target_rate * 100, 4),
        "data_note": data_note,
        "out": str(out),
        "size": out.stat().st_size,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
